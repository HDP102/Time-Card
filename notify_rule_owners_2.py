"""
notify_rule_owners.py
=====================
Workstream 1 — Rule Recertification Owner Notifications
PG&E | NPS Automation

What this script does:
  1. Reads the firewall rule spreadsheet (all device tabs)
  2. Parses each rule's AMPS owner data from Source/Destination columns
  3. Deduplicates owners across rules — one notification per owner
  4. Sends each owner a notification with three options:
       A) Recertify  B) Clean up / Remove  C) Review with the team
  5. Rules <= ATTACHMENT_THRESHOLD: inline table in email body with Decision column
     Rules >  ATTACHMENT_THRESHOLD: Excel attachment with Decision dropdown + note in body
  6. Supports Email (SMTP) and/or Teams (incoming webhook)
  7. Logs all notifications sent to a CSV for tracking

Usage:
  py -m pip install openpyxl requests
  py notify_rule_owners.py

Configuration (edit CONFIG block below):
  EXCEL_PATH           — path to the firewall rule spreadsheet
  NOTIFY_EMAIL         — True/False
  NOTIFY_TEAMS         — True/False
  DRY_RUN              — True = print only, do NOT send anything
  TEST_MODE            — True = send 2 test emails to yourself (one >8 rules, one <=8 rules)
  SENDER_EMAIL         — your CorpID@pge.com address
  SMTP_HOST            — mailhost (PG&E internal)
  TEAMS_WEBHOOK        — your Teams incoming webhook URL
  ATTACHMENT_THRESHOLD — rules above this number get an Excel attachment instead of inline table
"""

import ast
import csv
import io
import json
import logging
import smtplib
import re
from datetime import datetime, timezone
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from pathlib import Path

import requests
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
EXCEL_PATH           = "firewall_rule_report.xlsx"  # Path to the firewall rule spreadsheet
NOTIFY_EMAIL         = True                          # Send email notifications
NOTIFY_TEAMS         = False                         # Send Teams notifications
DRY_RUN              = True                          # True = print only, don't send

SENDER_EMAIL         = "corpid@pge.com"              # Your CorpID@pge.com
SMTP_HOST            = "mailhost"                    # PG&E internal mailhost
SMTP_PORT            = 25

TEAMS_WEBHOOK        = "<YOUR_TEAMS_WEBHOOK_URL>"   # Teams incoming webhook URL

LOG_FILE             = f"notifications_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

# TEST MODE — sends 2 emails to SENDER_EMAIL (yourself):
#   1. First owner with MORE than ATTACHMENT_THRESHOLD rules (tests attachment format)
#   2. First owner with ATTACHMENT_THRESHOLD or fewer rules (tests inline table format)
# Set DRY_RUN = False and TEST_MODE = True to run.
TEST_MODE            = False

# Rules above this number get an Excel attachment instead of an inline table
ATTACHMENT_THRESHOLD = 8

# Tags to SKIP — base rules and standard exceptions don't need owner notifications
SKIP_TAGS            = {"BaseRule", "ToolsRule", "ToolsRules"}

# Shadowing: skip fully shadowed rules (already handled by decommission workstream)
SKIP_SHADOWED        = True
# ─────────────────────────────────────────────

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

DECISION_OPTIONS = "A) Recertify  |  B) Clean up / Remove  |  C) Review with team"


# ── Owner parsing ──────────────────────────────────────────────────────────────

def parse_owner_string(raw):
    owners = []
    if not raw or "Skip AMPS" in raw or "No APPID" in raw:
        return owners
    for line in raw.strip().split("\n"):
        line = line.strip()
        if not line or "=" not in line:
            continue
        try:
            ip_part, dict_part = line.split("=", 1)
            ip = ip_part.strip()
            data = ast.literal_eval(dict_part.strip())
            for app_id, hostname_map in data.items():
                for hostname, inner in hostname_map.items():
                    for app_key, fields in inner.items():
                        if not isinstance(fields, dict):
                            continue
                        for role, value in fields.items():
                            if not value or not isinstance(value, str):
                                continue
                            match = re.match(r"(.+?)\s*\(([A-Z0-9]+)\)", value)
                            if match:
                                name   = match.group(1).strip()
                                corpid = match.group(2).strip()
                                email  = f"{corpid}@pge.com"
                                owners.append({
                                    "email":    email,
                                    "name":     name,
                                    "corpid":   corpid,
                                    "role":     role,
                                    "app_id":   app_id,
                                    "ip":       ip,
                                    "hostname": hostname,
                                })
        except Exception:
            continue
    return owners


def extract_primary_owner(owners):
    priority = ["Client Owner", "Cyber Owner", "IT SME", "IT Lead", "IT SME backup"]
    for role in priority:
        for o in owners:
            if o["role"] == role:
                return o
    return owners[0] if owners else None


# ── Excel reading ──────────────────────────────────────────────────────────────

def load_rules(excel_path):
    log.info(f"Loading spreadsheet: {excel_path}")
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    rules = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = None

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = list(row)
                col = {h: idx for idx, h in enumerate(headers) if h}
                continue
            if not any(row):
                continue

            def g(field):
                idx = col.get(field)
                return row[idx] if idx is not None else None

            tags = str(g("Tags") or "")
            if any(t in tags for t in SKIP_TAGS):
                continue

            shadow = str(g("Shadowing Status") or "")
            if SKIP_SHADOWED and shadow == "FULLY_SHADOWED":
                continue

            rules.append({
                "sheet":        sheet_name,
                "device_name":  g("Device Name"),
                "policy_name":  g("Policy Name"),
                "rule_name":    g("ID on Device"),
                "source":       g("Source"),
                "destination":  g("Destination"),
                "service":      g("Service"),
                "application":  g("Application"),
                "action":       g("Action"),
                "tags":         tags,
                "last_hit":     g("Last Hit"),
                "shadowing":    shadow,
                "ticket_id":    g("Ticket ID"),
                "app_name":     g("Application Name"),
                "disabled":     g("Disabled"),
                "src_owners_raw": g("Source AMPS owners list"),
                "dst_owners_raw": g("Destination AMPS owners list"),
                "business_justification": g("Policy Name + tag + FER close date + Business Justification"),
            })

    total_sheets = len(wb.sheetnames)
    wb.close()
    log.info(f"Loaded {len(rules)} rules across {total_sheets} devices")
    return rules


# ── Owner aggregation ──────────────────────────────────────────────────────────

def build_owner_rule_map(rules):
    owner_map = {}
    for rule in rules:
        all_owners = []
        all_owners += parse_owner_string(rule.get("src_owners_raw", "") or "")
        all_owners += parse_owner_string(rule.get("dst_owners_raw", "") or "")
        owner = extract_primary_owner(all_owners)
        if not owner:
            continue
        email = owner["email"]
        if email not in owner_map:
            owner_map[email] = {"owner": owner, "rules": []}
        owner_map[email]["rules"].append(rule)
    return owner_map


# ── Ticket ID cleaning ─────────────────────────────────────────────────────────

def clean_ticket_id(raw_ticket):
    ticket_lines = []
    for t in str(raw_ticket or "").strip().split("\n"):
        t = t.strip()
        if t and t not in ("0", "0.0"):
            try:
                ticket_lines.append(str(int(float(t))) if "." in t else t)
            except ValueError:
                ticket_lines.append(t)
    return ", ".join(ticket_lines) if ticket_lines else "N/A"


def truncate(text, length=60):
    text = str(text or "")
    return (text[:length] + "...") if len(text) > length else text


# ── Excel attachment builder ───────────────────────────────────────────────────

def build_excel_attachment(owner, rules):
    """Build an Excel file in memory with a Decision dropdown column for the owner to fill in."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Rules for Review"

    blue_fill  = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    light_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # Row 1 — instruction banner
    ws.merge_cells("A1:G1")
    ws["A1"] = "Firewall Rule Recertification — Please fill in the Decision column for each rule using the dropdown, then reply with this completed file."
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=11)
    ws["A1"].fill = blue_fill
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32

    # Row 2 — A/B/C legend
    ws.merge_cells("A2:G2")
    ws["A2"] = (
        "A) Recertify — Rule is still needed and should remain active     |     "
        "B) Clean up / Remove — Rule is no longer needed and can be disabled/deleted     |     "
        "C) Review with team — Unsure, would like to discuss with the NPS Automation team"
    )
    ws["A2"].font = Font(bold=True, color="003366", size=10)
    ws["A2"].fill = light_fill
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 28

    # Row 3 — column headers
    headers = ["Device", "Rule Name", "Source", "Destination", "Last Hit", "Ticket ID", "Decision (A / B / C)"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = blue_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="left")

    # Data rows starting at row 4
    for row_idx, r in enumerate(rules, 4):
        ws.cell(row=row_idx, column=1, value=str(r.get("device_name") or ""))
        ws.cell(row=row_idx, column=2, value=str(r.get("rule_name") or ""))
        ws.cell(row=row_idx, column=3, value=str(r.get("source") or ""))
        ws.cell(row=row_idx, column=4, value=str(r.get("destination") or ""))
        ws.cell(row=row_idx, column=5, value=str(r.get("last_hit") or "Never"))
        ws.cell(row=row_idx, column=6, value=clean_ticket_id(r.get("ticket_id")))
        cell = ws.cell(row=row_idx, column=7, value="")
        cell.alignment = Alignment(horizontal="center")

    # Dropdown on Decision column — rows 4 to end of data
    last_data_row = len(rules) + 3
    dv = DataValidation(
        type="list",
        formula1='"A) Recertify,B) Clean up / Remove,C) Review with team"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.sqref = f"G4:G{last_data_row}"
    ws.add_data_validation(dv)

    # Column widths
    col_widths = [25, 30, 40, 40, 15, 15, 28]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


# ── Email formatting ───────────────────────────────────────────────────────────

def get_first_name(owner):
    parts = owner["name"].split(",")
    return parts[1].strip().split()[0] if len(parts) > 1 else parts[0].strip()


def format_email_inline(owner, rules):
    """Email with inline rule table (8 or fewer rules)."""
    name = get_first_name(owner)

    rows = ""
    for r in rules:
        last_hit = str(r.get("last_hit") or "Never")
        ticket_display = clean_ticket_id(r.get("ticket_id"))
        src = truncate(r.get("source"))
        dst = truncate(r.get("destination"))
        rows += f"""
        <tr>
          <td style='padding:6px;border:1px solid #ddd'>{r['device_name'] or ''}</td>
          <td style='padding:6px;border:1px solid #ddd'>{r['rule_name'] or ''}</td>
          <td style='padding:6px;border:1px solid #ddd;font-size:11px'>{src}</td>
          <td style='padding:6px;border:1px solid #ddd;font-size:11px'>{dst}</td>
          <td style='padding:6px;border:1px solid #ddd'>{last_hit}</td>
          <td style='padding:6px;border:1px solid #ddd'>{ticket_display}</td>
          <td style='padding:6px;border:1px solid #ddd;text-align:center;color:#555;font-size:11px'>A / B / C</td>
        </tr>"""

    return f"""
    <html><body style='font-family:Arial,sans-serif;color:#333'>
    <p>Hi {name},</p>
    <p>As part of PG&E's annual firewall rule recertification process, the following
    firewall rules have been identified as requiring your review. These rules are
    associated with applications or assets you own.</p>
    <p><strong>Please review the rules below and reply to this email with your decision
    for each rule:</strong></p>
    <ol>
      <li><strong>A) Recertify</strong> — The rule is still needed and should remain active.</li>
      <li><strong>B) Clean up / Remove</strong> — The rule is no longer needed and can be disabled/deleted.</li>
      <li><strong>C) Review with the team</strong> — You are unsure and would like to discuss with the NPS Automation team.</li>
    </ol>
    <p>Enter your decision (A, B, or C) in the <strong>Decision</strong> column when replying.</p>
    <table style='border-collapse:collapse;width:100%;font-size:13px'>
      <thead>
        <tr style='background:#003366;color:white'>
          <th style='padding:8px;text-align:left'>Device</th>
          <th style='padding:8px;text-align:left'>Rule Name</th>
          <th style='padding:8px;text-align:left'>Source</th>
          <th style='padding:8px;text-align:left'>Destination</th>
          <th style='padding:8px;text-align:left'>Last Hit</th>
          <th style='padding:8px;text-align:left'>Ticket ID</th>
          <th style='padding:8px;text-align:center'>Decision (A / B / C)</th>
        </tr>
      </thead>
      <tbody>{rows}</tbody>
    </table>
    <p style='margin-top:20px'>Please reply to this email with your completed decisions,
    or contact the NPS Automation team to schedule a review.</p>
    <p>If you have any questions, please reach out directly.</p>
    <p>Thank you,<br><strong>NPS Automation Team</strong><br>PG&E Network &amp; Platform Security</p>
    <p style='font-size:11px;color:#888'>This is an automated notification from the NPS firewall recertification process.
    Rule count in this notification: {len(rules)}</p>
    </body></html>
    """


def format_email_attachment(owner, rules):
    """Email body when rules are in an attachment (more than 8 rules)."""
    name = get_first_name(owner)

    return f"""
    <html><body style='font-family:Arial,sans-serif;color:#333'>
    <p>Hi {name},</p>
    <p>As part of PG&E's annual firewall rule recertification process, the following
    firewall rules have been identified as requiring your review. These rules are
    associated with applications or assets you own.</p>
    <p>Due to the number of rules associated with your assets (<strong>{len(rules)} rules</strong>),
    your full rule list has been attached as an Excel file.</p>
    <p><strong>Please open the attached file, fill in the Decision column for each rule,
    and reply to this email with the completed attachment:</strong></p>
    <ol>
      <li><strong>A) Recertify</strong> — The rule is still needed and should remain active.</li>
      <li><strong>B) Clean up / Remove</strong> — The rule is no longer needed and can be disabled/deleted.</li>
      <li><strong>C) Review with the team</strong> — You are unsure and would like to discuss with the NPS Automation team.</li>
    </ol>
    <p>If you have any questions or would like to schedule a review session,
    please contact the NPS Automation team directly.</p>
    <p>Thank you,<br><strong>NPS Automation Team</strong><br>PG&E Network &amp; Platform Security</p>
    <p style='font-size:11px;color:#888'>This is an automated notification from the NPS firewall recertification process.
    Rule count in this notification: {len(rules)}</p>
    </body></html>
    """


# ── Sending ────────────────────────────────────────────────────────────────────

def send_email(to_email, owner, rules, dry_run=True):
    use_attachment = len(rules) > ATTACHMENT_THRESHOLD
    rule_count     = len(rules)
    subject        = f"Action Required: Firewall Rule Recertification ({rule_count} rule{'s' if rule_count > 1 else ''})"

    if use_attachment:
        html_body  = format_email_attachment(owner, rules)
        attach_format = "with Excel attachment"
    else:
        html_body  = format_email_inline(owner, rules)
        attach_format = "inline table"

    if dry_run:
        log.info(f"[DRY RUN] Would email: {to_email} | {rule_count} rules | {attach_format} | Subject: {subject}")
        return True

    msg = MIMEMultipart("mixed")
    msg["Subject"] = subject
    msg["From"]    = SENDER_EMAIL
    msg["To"]      = to_email
    msg.attach(MIMEText(html_body, "html"))

    if use_attachment:
        excel_bytes = build_excel_attachment(owner, rules)
        part = MIMEBase("application", "octet-stream")
        part.set_payload(excel_bytes)
        encoders.encode_base64(part)
        filename = f"Firewall_Rules_Recertification_{owner['corpid']}.xlsx"
        part.add_header("Content-Disposition", f"attachment; filename={filename}")
        msg.attach(part)

    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.sendmail(SENDER_EMAIL, [to_email], msg.as_string())
        log.info(f"Email sent: {to_email} | {rule_count} rules | {attach_format}")
        return True
    except Exception as e:
        log.error(f"Email failed to {to_email}: {e}")
        return False


def send_teams(owner, rules, dry_run=True):
    name = get_first_name(owner)
    use_attachment = len(rules) > ATTACHMENT_THRESHOLD

    rule_lines = []
    for r in rules:
        last_hit = str(r.get("last_hit") or "Never")
        ticket_display = clean_ticket_id(r.get("ticket_id"))
        rule_lines.append(
            f"• **{r['rule_name'] or 'Unnamed'}** | Device: {r['device_name']} | "
            f"Last Hit: {last_hit} | Ticket: {ticket_display}"
        )

    rules_text = "\n\n".join(rule_lines)
    if use_attachment:
        rules_text = f"⚠️ **{len(rules)} rules** — see the email attachment for the full list."

    payload = {
        "type": "message",
        "attachments": [{
            "contentType": "application/vnd.microsoft.card.adaptive",
            "content": {
                "$schema": "http://adaptivecards.io/schemas/adaptive-card.json",
                "type": "AdaptiveCard",
                "version": "1.4",
                "body": [
                    {"type": "TextBlock", "text": "🔒 Firewall Rule Recertification Required",
                     "weight": "Bolder", "size": "Medium", "color": "Accent"},
                    {"type": "TextBlock",
                     "text": f"Hi **{name}**, the following firewall rules require your review.",
                     "wrap": True},
                    {"type": "TextBlock", "text": "**Your Rules:**", "weight": "Bolder", "spacing": "Medium"},
                    {"type": "TextBlock", "text": rules_text, "wrap": True, "spacing": "Small"},
                    {"type": "TextBlock", "text": "**Decision Options:**", "weight": "Bolder", "spacing": "Medium"},
                    {"type": "TextBlock",
                     "text": "✅ **A) Recertify** — Rule is still needed\n\n"
                             "🗑️ **B) Clean up / Remove** — Rule is no longer needed\n\n"
                             "💬 **C) Review with the team** — Need to discuss",
                     "wrap": True, "spacing": "Small"},
                    {"type": "TextBlock",
                     "text": "Reply to the email with your decisions, or contact the NPS Automation team.",
                     "wrap": True, "spacing": "Medium", "isSubtle": True}
                ]
            }
        }]
    }

    if dry_run:
        log.info(f"[DRY RUN] Would Teams-notify: {owner['name']} | {len(rules)} rules")
        return True

    if not TEAMS_WEBHOOK or TEAMS_WEBHOOK.startswith("<"):
        log.warning("Teams webhook not configured, skipping")
        return False

    try:
        resp = requests.post(TEAMS_WEBHOOK, headers={"Content-Type": "application/json"},
                             data=json.dumps(payload), timeout=10)
        if resp.status_code in (200, 202):
            log.info(f"Teams notification sent: {owner['name']}")
            return True
        else:
            log.error(f"Teams failed for {owner['name']}: {resp.status_code} {resp.text}")
            return False
    except Exception as e:
        log.error(f"Teams error for {owner['name']}: {e}")
        return False


# ── Logging ────────────────────────────────────────────────────────────────────

def write_log(log_file, rows):
    with open(log_file, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "timestamp", "email", "name", "corpid", "role",
            "rule_count", "attachment_sent", "email_sent", "teams_sent", "device_list"
        ])
        writer.writeheader()
        writer.writerows(rows)
    log.info(f"Notification log written to: {log_file}")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    if DRY_RUN:
        log.info("=" * 60)
        log.info("DRY RUN MODE — no notifications will actually be sent")
        log.info("Set DRY_RUN = False in CONFIG to send for real")
        log.info("=" * 60)

    rules     = load_rules(EXCEL_PATH)
    owner_map = build_owner_rule_map(rules)
    log.info(f"Unique owners to notify: {len(owner_map)}")

    # ── TEST MODE: send one >threshold and one <=threshold email to yourself
    if TEST_MODE and not DRY_RUN:
        log.info("=" * 60)
        log.info("TEST MODE — sending 2 sample emails to yourself")
        log.info(f"Recipient overridden to: {SENDER_EMAIL}")
        log.info("=" * 60)

        large_owner = next(
            (d for d in owner_map.values() if len(d["rules"]) > ATTACHMENT_THRESHOLD), None
        )
        small_owner = next(
            (d for d in owner_map.values() if len(d["rules"]) <= ATTACHMENT_THRESHOLD), None
        )

        if large_owner:
            log.info(f"Test 1: Sending ATTACHMENT email ({len(large_owner['rules'])} rules)")
            send_email(SENDER_EMAIL, large_owner["owner"], large_owner["rules"], dry_run=False)
        else:
            log.warning("No owner with more than threshold rules found for attachment test.")

        if small_owner:
            log.info(f"Test 2: Sending INLINE email ({len(small_owner['rules'])} rules)")
            send_email(SENDER_EMAIL, small_owner["owner"], small_owner["rules"], dry_run=False)
        else:
            log.warning("No owner with threshold or fewer rules found for inline test.")

        log.info("Test emails sent. Review your inbox then set TEST_MODE = False to run for real.")
        return

    # ── Production / dry run loop
    log_rows = []
    for email, data in owner_map.items():
        owner       = data["owner"]
        owner_rules = data["rules"]
        email_ok    = False
        teams_ok    = False

        if NOTIFY_EMAIL:
            email_ok = send_email(email, owner, owner_rules, dry_run=DRY_RUN)
        if NOTIFY_TEAMS:
            teams_ok = send_teams(owner, owner_rules, dry_run=DRY_RUN)

        devices = list({r["device_name"] for r in owner_rules if r["device_name"]})
        log_rows.append({
            "timestamp":       datetime.now(timezone.utc).isoformat(),
            "email":           email,
            "name":            owner["name"],
            "corpid":          owner["corpid"],
            "role":            owner["role"],
            "rule_count":      len(owner_rules),
            "attachment_sent": len(owner_rules) > ATTACHMENT_THRESHOLD,
            "email_sent":      email_ok,
            "teams_sent":      teams_ok,
            "device_list":     "; ".join(devices[:5]) + ("..." if len(devices) > 5 else ""),
        })

    write_log(LOG_FILE, log_rows)

    # Summary
    attached_count = sum(1 for r in log_rows if r["attachment_sent"])
    inline_count   = len(log_rows) - attached_count

    print("\n" + "=" * 60)
    print("  NOTIFICATION SUMMARY")
    print(f"  Total rules processed:    {len(rules)}")
    print(f"  Unique owners:            {len(owner_map)}")
    print(f"  Inline table emails:      {inline_count}")
    print(f"  Attachment emails:        {attached_count}")
    print(f"  Email notifications:      {'enabled' if NOTIFY_EMAIL else 'disabled'}")
    print(f"  Teams notifications:      {'enabled' if NOTIFY_TEAMS else 'disabled'}")
    print(f"  Dry run:                  {DRY_RUN}")
    print(f"  Log file:                 {LOG_FILE}")
    print("=" * 60)


if __name__ == "__main__":
    main()
