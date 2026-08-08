"""
ZIA Custom URL Category Extractor
------------------------------------
Extracts custom URL categories from a Zscaler admin portal HAR file
and cross-references them against URL filtering and SSL inspection
policies to build a complete inventory.

PowerShell:
    pip install openpyxl
    python zia_category_extractor.py Zscaler.har url_filtering.json sslpol.json

Output (in current directory):
    - ZIA_Web_Category_Inventory.xlsx
"""

import json
import sys
import os
import re

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)


# ============================================================
# PARSING
# ============================================================

def parse_har(har_path):
    """Extract custom URL categories from HAR file."""
    print(f"  Parsing HAR: {har_path}")
    with open(har_path, "r", encoding="utf-8") as f:
        har = json.load(f)

    categories = []
    seen = set()
    for entry in har["log"]["entries"]:
        url = entry["request"]["url"]
        if "urlCategories/CUSTOM_" in url:
            text = entry["response"].get("content", {}).get("text", "")
            if text:
                try:
                    data = json.loads(text)
                    cat_id = data.get("id", "")
                    if cat_id and cat_id not in seen:
                        seen.add(cat_id)
                        categories.append(data)
                except (json.JSONDecodeError, KeyError):
                    pass

    print(f"  Found {len(categories)} custom categories")
    return categories


def build_usage_maps(url_path, ssl_path):
    """Build maps of which rules reference which custom categories."""
    url_usage = {}
    ssl_usage = {}
    std_cats = set()

    with open(url_path, "r", encoding="utf-8") as f:
        url_rules = json.load(f)
    with open(ssl_path, "r", encoding="utf-8") as f:
        ssl_rules = json.load(f)

    for r in url_rules:
        for cat in r.get("urlCategories", []):
            if cat.startswith("CUSTOM_"):
                url_usage.setdefault(cat, []).append(r["name"])
            else:
                std_cats.add(cat)

    for r in ssl_rules:
        for cat in r.get("urlCategories", []):
            if cat.startswith("CUSTOM_"):
                ssl_usage.setdefault(cat, []).append(r["name"])

    print(f"  URL filtering: {len(url_usage)} custom categories referenced")
    print(f"  SSL inspection: {len(ssl_usage)} custom categories referenced")
    print(f"  Standard web categories: {len(std_cats)}")

    return url_usage, ssl_usage, sorted(std_cats)


# ============================================================
# EXCEL BUILDING
# ============================================================

BOLD = Font(name="Calibri", bold=True, size=11)
NORMAL = Font(name="Calibri", size=11)
SMALL_ITAL = Font(name="Calibri", size=10, italic=True)
TITLE = Font(name="Calibri", bold=True, size=14)
SUBTITLE = Font(name="Calibri", bold=True, size=11)
HEADER_FILL = PatternFill("solid", fgColor="404040")
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
LIGHT_RED = PatternFill("solid", fgColor="F2DCDB")
BORDER = Border(
    left=Side(style="thin", color="AAAAAA"),
    right=Side(style="thin", color="AAAAAA"),
    top=Side(style="thin", color="AAAAAA"),
    bottom=Side(style="thin", color="AAAAAA"),
)


def hdr(ws, row, headers):
    for c, h in enumerate(headers, 1):
        cl = ws.cell(row=row, column=c, value=h)
        cl.font = HEADER_FONT
        cl.fill = HEADER_FILL
        cl.border = BORDER
        cl.alignment = Alignment(wrap_text=True, vertical="center")


def cel(ws, row, col, val, font=None, fill=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = font or NORMAL
    c.border = BORDER
    if fill:
        c.fill = fill
    c.alignment = Alignment(wrap_text=True, vertical="top")
    return c


def truncate_list(items, limit, sep=", "):
    """Join list items, truncating with count if over limit."""
    if not items:
        return ""
    shown = sep.join(items[:limit])
    if len(items) > limit:
        shown += f"... (+{len(items) - limit} more)"
    return shown


def build_workbook(categories, url_usage, ssl_usage, std_cats):
    wb = Workbook()

    # Prepare category data
    cat_data = []
    for c in sorted(categories, key=lambda x: int(x["id"].replace("CUSTOM_", ""))):
        cat_id = c["id"]
        urls = c.get("urls", []) + c.get("dbCategorizedUrls", [])
        cat_data.append({
            "id": cat_id,
            "name": c.get("configuredName", ""),
            "super": c.get("superCategory", ""),
            "urls": urls,
            "url_count": len(urls),
            "url_rules": url_usage.get(cat_id, []),
            "ssl_rules": ssl_usage.get(cat_id, []),
        })

    # ---- TAB 1: Summary ----
    ws1 = wb.active
    ws1.title = "Summary"
    ws1["A1"] = "ZIA Web Categories & Custom URL Categories"
    ws1["A1"].font = TITLE

    ws1["A3"] = "Custom URL Categories"
    ws1["A3"].font = SUBTITLE
    stats = [
        ("Total custom categories", len(cat_data)),
        ("Used in URL filtering", sum(1 for c in cat_data if c["url_rules"])),
        ("Used in SSL inspection", sum(1 for c in cat_data if c["ssl_rules"])),
        ("Used in both policies", sum(1 for c in cat_data if c["url_rules"] and c["ssl_rules"])),
        ("Not used in either policy", sum(1 for c in cat_data if not c["url_rules"] and not c["ssl_rules"])),
        ("", ""),
        ("With URLs/domains defined", sum(1 for c in cat_data if c["urls"])),
        ("Empty (no URLs)", sum(1 for c in cat_data if not c["urls"])),
    ]
    r = 4
    for label, val in stats:
        if label:
            ws1.cell(row=r, column=1, value=label).font = NORMAL
            ws1.cell(row=r, column=2, value=val).font = BOLD
        r += 1

    r += 1
    ws1.cell(row=r, column=1, value="Standard Web Categories").font = SUBTITLE
    r += 1
    ws1.cell(row=r, column=1, value="Predefined ZIA categories referenced in URL filtering").font = NORMAL
    ws1.cell(row=r, column=2, value=len(std_cats)).font = BOLD

    r += 2
    ws1.cell(row=r, column=1, value="Nomenclature").font = SUBTITLE
    r += 1
    terms = [
        ("Standard Web Categories", "Zscaler-maintained (~110 built-in). STREAMING_MEDIA, SOCIAL_NETWORKING, etc. Zscaler auto-categorizes every URL."),
        ("Custom URL Categories", "Created internally. Contain specific URLs/domains. Override Zscaler built-in categorization."),
        ("Cloud App Control", "Separate policy controlling cloud applications (Facebook, YouTube). Different from URL categories."),
    ]
    for term, defn in terms:
        ws1.cell(row=r, column=1, value=term).font = BOLD
        ws1.cell(row=r, column=2, value=defn).font = NORMAL
        ws1.merge_cells(f"B{r}:D{r}")
        ws1.cell(row=r, column=2).alignment = Alignment(wrap_text=True)
        ws1.row_dimensions[r].height = 35
        r += 1

    ws1.column_dimensions["A"].width = 38
    ws1.column_dimensions["B"].width = 50

    # ---- TAB 2: Custom URL Categories ----
    ws2 = wb.create_sheet("Custom URL Categories")
    hdr(ws2, 1, ["Category ID", "Name", "URL Count", "URLs / Domains", "URL Filtering Rules", "SSL Inspection Rules", "Used In"])

    r = 2
    for cat in cat_data:
        cel(ws2, r, 1, cat["id"])
        cel(ws2, r, 2, cat["name"])
        cel(ws2, r, 3, cat["url_count"])
        cel(ws2, r, 4, truncate_list(cat["urls"], 20))
        cel(ws2, r, 5, truncate_list(cat["url_rules"], 5))
        cel(ws2, r, 6, truncate_list(cat["ssl_rules"], 5))

        if cat["url_rules"] and cat["ssl_rules"]:
            cel(ws2, r, 7, "Both")
        elif cat["url_rules"]:
            cel(ws2, r, 7, "URL only")
        elif cat["ssl_rules"]:
            cel(ws2, r, 7, "SSL only")
        else:
            cel(ws2, r, 7, "UNUSED", fill=LIGHT_RED)
        r += 1

    ws2.auto_filter.ref = f"A1:G{r - 1}"
    ws2.freeze_panes = "A2"
    for i, w in enumerate([14, 35, 10, 60, 45, 45, 12], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ---- TAB 3: Unused Categories ----
    ws3 = wb.create_sheet("Unused Categories")
    ws3["A1"] = "Custom categories not referenced in any URL filtering or SSL inspection rule"
    ws3["A1"].font = SUBTITLE
    ws3.merge_cells("A1:D1")

    hdr(ws3, 2, ["Category ID", "Name", "URL Count", "URLs / Domains"])
    r = 3
    for cat in cat_data:
        if not cat["url_rules"] and not cat["ssl_rules"]:
            cel(ws3, r, 1, cat["id"])
            cel(ws3, r, 2, cat["name"])
            cel(ws3, r, 3, cat["url_count"])
            cel(ws3, r, 4, truncate_list(cat["urls"], 15))
            r += 1

    for i, w in enumerate([14, 35, 10, 60], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = "A3"

    # ---- TAB 4: Cross-Policy Usage ----
    ws4 = wb.create_sheet("Cross-Policy Usage")
    ws4["A1"] = "Categories used in BOTH URL filtering and SSL inspection"
    ws4["A1"].font = SUBTITLE
    ws4.merge_cells("A1:E1")

    hdr(ws4, 2, ["Category ID", "Name", "URL Filtering Rule(s)", "SSL Inspection Rule(s)", "Mismatch?"])
    r = 3
    for cat in cat_data:
        if cat["url_rules"] and cat["ssl_rules"]:
            cel(ws4, r, 1, cat["id"])
            cel(ws4, r, 2, cat["name"])
            cel(ws4, r, 3, ", ".join(cat["url_rules"]))
            cel(ws4, r, 4, ", ".join(cat["ssl_rules"]))

            # Check for obvious mismatches: different service names in rules
            url_names = " ".join(cat["url_rules"]).lower()
            ssl_names = " ".join(cat["ssl_rules"]).lower()
            cat_name = cat["name"].lower().replace("_", " ")

            # Flag if category name keywords don't appear in one of the policy rule names
            cat_words = [w for w in cat_name.split() if len(w) > 3]
            url_has = any(w in url_names for w in cat_words) if cat_words else True
            ssl_has = any(w in ssl_names for w in cat_words) if cat_words else True

            if not url_has or not ssl_has:
                cel(ws4, r, 5, "Review — rule names may reference different services", fill=LIGHT_RED)
            else:
                cel(ws4, r, 5, "")

            ws4.row_dimensions[r].height = 30
            r += 1

    for i, w in enumerate([14, 30, 40, 45, 40], 1):
        ws4.column_dimensions[get_column_letter(i)].width = w
    ws4.freeze_panes = "A3"

    # ---- TAB 5: Standard Web Categories ----
    ws5 = wb.create_sheet("Standard Web Categories")
    ws5["A1"] = "Predefined ZIA web categories referenced in URL filtering rules"
    ws5["A1"].font = SUBTITLE
    ws5.merge_cells("A1:B1")

    hdr(ws5, 2, ["#", "Category Name"])
    for i, cat in enumerate(sorted(std_cats), 1):
        cel(ws5, i + 2, 1, i)
        cel(ws5, i + 2, 2, cat)

    ws5.column_dimensions["A"].width = 6
    ws5.column_dimensions["B"].width = 45

    return wb


# ============================================================
# MAIN
# ============================================================

def main():
    if len(sys.argv) < 4:
        print(__doc__)
        sys.exit(1)

    har_path = sys.argv[1]
    url_path = sys.argv[2]
    ssl_path = sys.argv[3]

    for path in [har_path, url_path, ssl_path]:
        if not os.path.exists(path):
            print(f"ERROR: File not found: {path}")
            sys.exit(1)

    print("Step 1: Extracting custom categories from HAR...")
    categories = parse_har(har_path)

    if not categories:
        print("WARNING: No custom categories found in HAR file.")
        print("Make sure you expanded all categories and scrolled through")
        print("the URL Categories page before exporting the HAR.")

    print("\nStep 2: Building policy cross-reference...")
    url_usage, ssl_usage, std_cats = build_usage_maps(url_path, ssl_path)

    print("\nStep 3: Building Excel workbook...")
    wb = build_workbook(categories, url_usage, ssl_usage, std_cats)

    output = os.path.join(os.getcwd(), "ZIA_Web_Category_Inventory.xlsx")
    wb.save(output)
    print(f"\nSaved: {output}")

    # Print summary
    used_url = sum(1 for c in categories if c["id"] in url_usage)
    used_ssl = sum(1 for c in categories if c["id"] in ssl_usage)
    unused = sum(1 for c in categories if c["id"] not in url_usage and c["id"] not in ssl_usage)
    print(f"\n  {len(categories)} custom categories")
    print(f"  {used_url} used in URL filtering")
    print(f"  {used_ssl} used in SSL inspection")
    print(f"  {unused} unused (cleanup candidates)")
    print(f"  {len(std_cats)} standard web categories")
    print("\nDone!")


if __name__ == "__main__":
    main()
