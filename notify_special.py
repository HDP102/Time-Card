"""
notify_rule_owners.py
=====================
Workstream 1 — Rule Recertification Owner Notifications
PG&E | NPS Automation

Reads the firewall rule spreadsheet, resolves each rule's owner from the AMPS data,
groups rules by owner, and contacts them with three options:
  A) Recertify   B) Clean up / Remove   C) Review with the team

Every owner receives an Excel attachment with a Decision dropdown — the response
format is uniform so replies can be tracked and audited consistently.

RUN MODES (set RUN_MODE in CONFIG):
  notify    Email each owner their rules, plus an optional companion Teams card.
  remind    Chase owners who have not responded, with a days-outstanding count.
  announce  One heads-up card to the team channel before a wave goes out.
  digest    One status card to the team channel. Contacts no owners.

Filters (NOTIFY_ONLY / EXCLUDE_OWNERS / NOTIFY_DEVICES) apply to every mode, and
in dry run, test, and real sends alike.

Teams notes:
  - Webhook URLs must come from the Workflows / Power Automate app. Classic Office
    365 connector webhooks were retired by Microsoft in May 2026.
  - Owner-directed cards need TEAMS_USER_WEBHOOK, a flow that reads a "recipient"
    field and DMs that person. Without it, owner cards fall back to the channel.
  - Cards never carry rule data — they point the owner at the email.

Usage:
  py -m pip install openpyxl requests
  py notify_rule_owners.py
"""

import ast
import csv
import io
import json
import logging
import smtplib
import re
import time
from datetime import datetime, timezone
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from pathlib import Path
from urllib.parse import quote

import requests
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────

# ── What this run does ───────────────────────────────────────────────
# RUN_MODE picks the job. Filters apply to every mode.
#   "notify"   — the main run. Emails each owner their rules as an Excel attachment,
#                and optionally posts a companion Teams card telling them to check email.
#   "remind"   — chases owners who have not responded yet. Same audience minus anyone
#                listed in RESPONDED_FILE, with reminder wording and a days-outstanding
#                count. Requires a prior notify run so there is history to read.
#   "announce" — one heads-up card to the team channel before a notify wave goes out.
#                Contacts no owners. Cuts down "is this email real?" replies.
#   "digest"   — one status card to the team channel: notified / responded / outstanding,
#                oldest outstanding, biggest holdouts. Contacts no owners.
RUN_MODE             = "notify"

# ════════════════════════════════════════════════════════════════════
#  BATCH PICKER  —  set BATCH once, run. No other per-run edits needed.
# ════════════════════════════════════════════════════════════════════
#   BATCH = 1   ->  GDN
#   BATCH = 2   ->  Medium
#   BATCH = 3   ->  ODN NERC Part 1
#   BATCH = 4   ->  ODN NERC Part 2
#   BATCH = 5   ->  ODN Remaining Part 1
#   BATCH = 6   ->  ODN Remaining Part 2
#   BATCH = 7   ->  ODN MPLS Part 1
#   BATCH = 8   ->  ALL of the above, one after another (hands-off)
#
BATCH = 1

# Each batch bundles everything that changes per list: the spreadsheet, the SharePoint
# list link, the subject scope, and the reporting-view export used to skip owners who
# have already responded. Update a link or filename here in ONE place.
BATCHES = {
    1: {
        "excel":     "GDN Rule Recertification.xlsx",
        "link":      "https://pge-my.sharepoint.com/:l:/r/personal/s9ch_pge_com/Lists/GDN%20Rule%20Recertification?e=ugBafO",
        "scope":     "GDN",
        "reporting": "GDN_reporting_view.xlsx",
    },
    2: {
        "excel":     "MediumRule_report_with_APPID.xlsx",
        "link":      "https://pge-my.sharepoint.com/:l:/r/personal/s9ch_pge_com/Lists/MediumRule_report_with_APPID?e=aYze5X",
        "scope":     "Medium",
        "reporting": "Medium_reporting_view.xlsx",
    },
    3: {
        "excel":     "ODN_NERC_Low_Medium_Part1.xlsx",
        "link":      "https://pge-my.sharepoint.com/:l:/r/personal/s9ch_pge_com/Lists/ODN_NERC_Low_Medium_Part1?e=H1amcw",
        "scope":     "ODN NERC Part 1",
        "reporting": "ODN_NERC_Part1_reporting_view.xlsx",
    },
    4: {
        "excel":     "ODN_NERC_Low_Medium_Part2.xlsx",
        "link":      "https://pge-my.sharepoint.com/:l:/r/personal/n1v3_pge_com/Lists/ODN_NERC_Low_Medium_Part2?e=58RSJF",
        "scope":     "ODN NERC Part 2",
        "reporting": "ODN_NERC_Part2_reporting_view.xlsx",
    },
    5: {
        "excel":     "Remaining_ODN_Firewalls_Part1.xlsx",
        "link":      "https://pge-my.sharepoint.com/:l:/r/personal/n1v3_pge_com/Lists/ODN_except_MPLS_Part1?e=TPiSNH",
        "scope":     "ODN Remaining Part 1",
        "reporting": "ODN_Remaining_Part1_reporting_view.xlsx",
    },
    6: {
        "excel":     "Remaining_ODN_Firewalls_Part2.xlsx",
        "link":      "https://pge-my.sharepoint.com/:l:/r/personal/n1v3_pge_com/Lists/Remaining_ODN_Firewalls_Part2?e=7bSjXf",
        "scope":     "ODN Remaining Part 2",
        "reporting": "ODN_Remaining_Part2_reporting_view.xlsx",
    },
    7: {
        "excel":     "ODN_MPLS_Rules_Part1.xlsx",
        "link":      "https://pge-my.sharepoint.com/:l:/r/personal/n1v3_pge_com/Lists/ODN_MPLS_Rules_Part1?e=ynna8p",
        "scope":     "ODN MPLS Part 1",
        "reporting": "ODN_MPLS_Part1_reporting_view.xlsx",
    },
}
ALL_BATCHES = 8   # BATCH value that runs every batch in sequence
# ════════════════════════════════════════════════════════════════════

EXCEL_PATH           = "firewall_rule_report.xlsx"  # Path to the firewall rule spreadsheet
NOTIFY_EMAIL         = True                          # Send email notifications
NOTIFY_TEAMS         = False                         # Send Teams notifications
DRY_RUN              = True                          # True = print only, don't send

SENDER_EMAIL         = "corpid@pge.com"              # Your CorpID@pge.com
SMTP_HOST            = "mailhost"                    # PG&E internal mailhost
SMTP_PORT            = 25

# ── Teams ────────────────────────────────────────────────────────────
# Webhook URLs must come from the Workflows / Power Automate app. Classic Office 365
# connector webhooks were retired by Microsoft in May 2026 and no longer deliver.
#
# TEAMS_WEBHOOK — the team channel. Used for "announce" and "digest" cards, and as the
#   fallback for owner cards if no per-user flow is configured.
TEAMS_WEBHOOK        = "<YOUR_TEAMS_CHANNEL_WEBHOOK>"

# TEAMS_USER_WEBHOOK — a Power Automate flow that DMs an individual. The flow must read
#   a "recipient" field from the request body and post to that person's chat. Owner cards
#   go here when set. Leave "" to fall back to the channel webhook above.
#   An owner's UPN is their CorpID address, which the script already has.
TEAMS_USER_WEBHOOK   = ""

# TEAMS_WEBHOOK_TEST — used instead of the above whenever TEST_MODE = True. Point it at
#   your own chat or a private channel. Leave "" to reuse the real destination.
TEAMS_WEBHOOK_TEST   = ""

# Seconds between Teams posts. Webhooks throttle a tight loop — keep at 1 or higher.
TEAMS_DELAY_SECONDS  = 1

LOG_FILE             = "notifications.csv"   # one running log; runs append to it

# Anchor shared paths to THIS script's location, not the working directory, so they
# resolve the same no matter where the script is launched or where OneDrive mounts.
SCRIPT_DIR           = Path(__file__).resolve().parent          # .../NPS-Automation/Notify
SHARED_ROOT          = SCRIPT_DIR.parent                        # .../NPS-Automation

# ── Response tracking (used by "remind" and "digest") ────────────────
# RESPONDED_FILE — CSV of owners who have already replied, so reminders skip them.
#   One CorpID per line, or a "corpid" column. Shared, so whoever processes replies
#   updates one file both tools see. Used as a fallback / manual override.
RESPONDED_FILE       = str(SHARED_ROOT / "responded.csv")

# SKIP_RESPONDED — when True, owners who have already recorded a decision are skipped,
#   so a run only emails people who still have blank Recertification entries. This is
#   what makes reminders / repeat sends only chase non-responders. Reads the reporting-
#   view export set in REPORTING_VIEW_PATH.
SKIP_RESPONDED       = True

# REPORTING_VIEW_PATH — the exported SharePoint "reporting view" for the list being sent
#   (the fast custom-view export). Its Recertification column tells us who has responded.
#   Any non-blank value there (Recertify / Recertify - DR / Cleanup - Remove / Need
#   Assistance) counts as responded. Set it per run to match the list you're sending.
REPORTING_VIEW_PATH  = "reporting_view.xlsx"

# RECERT_COLUMN — the column in the reporting view that holds each rule's decision.
RECERT_COLUMN        = "Recertification"

# RESPONSES_DIR — shared folder where replied-to spreadsheets are dropped. Sits at the
#   shared root so both this tool and the recert processor read the same place.
#   Created automatically after the first real send if it doesn't exist.
RESPONSES_DIR        = str(SHARED_ROOT / "responses")

# LOGS_DIR — run logs live here, under this tool's own folder (kept local, not shared,
#   so two people running at once don't collide on OneDrive). Reminders/digests read it.
#   To void a run, delete its rows from the log.
LOGS_DIR             = str(SCRIPT_DIR / "logs")

# REMIND_AFTER_DAYS — only chase owners notified at least this many days ago.
#   Notification dates are read from the notifications_*.csv logs of previous runs.
REMIND_AFTER_DAYS    = 14

# ── Owner roles (new role-column spreadsheet format) ─────────────────
# ROLE_COLUMNS — all owner-role columns that may appear in the sheet, read as
#   "Last, First (CORPID)". Order here is the order recipients are collected.
ROLE_COLUMNS = [
    "IT SME", "IT SME Backup", "IT Lead", "IT Lead Delegate",
    "Client Owner", "IT Director", "IT Senior Leader",
]

# NOTIFY_ROLES — which of those roles actually get notified. Per the review meetings:
#   IT SME, IT SME Backup, IT Lead confirmed; Client Owner likely (kept on, toggle off
#   if leadership says otherwise); IT Director / Senior Leader are NOT notified.
NOTIFY_ROLES = [
    "IT SME", "IT SME Backup", "IT Lead", "Client Owner",
]

# GROUP_BY — how notifications are grouped:
#   "owner"  — ONE email per person, covering everything they own across all app IDs.
#              This is the default: an owner on many apps gets a single email, not one
#              per app. Best for SharePoint mode, where the email just points people to
#              the list (there's no per-app content to lose by consolidating).
#   "app_id" — one email per app ID, sent to all NOTIFY_ROLES holders for that app. An
#              owner on many apps receives many emails. Only use this if you specifically
#              need per-app emails; at scale it floods heavy owners with duplicates.
GROUP_BY             = "owner"

# ── Response method ──────────────────────────────────────────────────
# RESPONSE_METHOD — how owners record their decisions:
#   "sharepoint" — email points to the SharePoint list; owners edit the
#                  recertification column there. No file to send back.
#   "attachment" — owners fill in the Excel attachment and reply with it.
# "sharepoint" still includes the attachment so owners can see their rules;
# it just directs them to record decisions in the list instead of replying.
RESPONSE_METHOD      = "sharepoint"      # "sharepoint" | "attachment"

# SHAREPOINT_LINK — the list owners are sent to when RESPONSE_METHOD = "sharepoint".
#   This differs per network (ODN vs GDN), so set it to match the sheet you're
#   sending. Use the list share link (the ".../:l:/r/..." form), not a single item.
#   Example (GDN):
#   https://pge-my.sharepoint.com/:l:/r/personal/<owner>/Lists/GDN%20Rule%20Recertification?e=XXXX
SHAREPOINT_LINK      = "<SHAREPOINT_LIST_LINK>"

# EMAIL_SCOPE — which list/network this run is for, shown in the subject so recipients
#   getting multiple emails can tell them apart (e.g. "GDN", "ODN", "ODN Part 2",
#   "Medium"). Set it per run alongside SHAREPOINT_LINK to match the sheet you're
#   sending. Leave "" to omit it from the subject.
EMAIL_SCOPE          = ""

# ── Email content (editable text shown to owners) ────────────────────
# The stable instructional prose (the ask, the four options, background, how-to,
# FAQ questions) lives in the email builder. The values below are the parts that
# change between runs or shouldn't be buried in code — edit them here.

# RECERT_DEADLINE — the hard "all rules must be recertified by" date shown in the email.
RECERT_DEADLINE      = "8/31"
# RECERT_WINDOW — the softer "please respond within" phrasing.
RECERT_WINDOW        = "one week"

# FAQ_LINK — the public How-To / FAQ document.
FAQ_LINK             = "https://pge.sharepoint.com/:w:/r/sites/NetworkProtectionServicesPublic/_layouts/15/Doc.aspx?sourcedoc=%7B63AB99B0-3F4D-4F8B-8AD2-5965EC76F1FE%7D&file=Firewall%20Recertification%20How-To%20and%20FAQ.docx&action=default&mobileredirect=true&DefaultItemOpen=1"

# OFFICE_HOURS_INFO — HTML shown in the office-hours section. Pre-filled with the
#   current office hours; update here if the meetings change (they carry live Teams
#   join links and passcodes).
OFFICE_HOURS_INFO    = (
    "<p><strong>AFTERNOON 2-3PM</strong> (Monday, Tuesday, Wednesday and Thursday)<br>"
    "Microsoft Teams meeting<br>"
    "Join: <a href=\"https://teams.microsoft.com/meet/237151648973835?p=oqRjctCrmJ53CH9QpF\">"
    "https://teams.microsoft.com/meet/237151648973835?p=oqRjctCrmJ53CH9QpF</a><br>"
    "Meeting ID: 237 151 648 973 835 &nbsp;|&nbsp; Passcode: e2fU6RY6<br>"
    "Dial in: +1 415-906-0873,,425270707# (San Francisco) &nbsp;|&nbsp; "
    "Phone conference ID: 425 270 707#</p>"
    "<p><strong>MORNING 10:30-11:30 AM</strong> (Monday, Tuesday, Wednesday, and Friday)<br>"
    "Microsoft Teams meeting<br>"
    "Join: <a href=\"https://teams.microsoft.com/meet/258618918930560?p=3bKX6QK8D2RcjLR1Wx\">"
    "https://teams.microsoft.com/meet/258618918930560?p=3bKX6QK8D2RcjLR1Wx</a><br>"
    "Meeting ID: 258 618 918 930 560 &nbsp;|&nbsp; Passcode: 3bJ2ZJ3Q<br>"
    "Dial in: +1 415-906-0873,,748003855# (San Francisco) &nbsp;|&nbsp; "
    "Phone conference ID: 748 003 855#</p>"
)

# SENDER_SIGNATURE — the sign-off shown at the bottom of the email. Pre-filled with
#   the current sender; change it to whoever (or whichever mailbox) is sending.
SENDER_SIGNATURE     = (
    "Megan Merritt<br>"
    "Manager, Network Protection Services<br>"
    "Cybersecurity Architecture, Engineering &amp; Operations<br>"
    "Pacific Gas and Electric Company<br>"
    "Megan.Merritt@pge.com<br>"
    "Desk: +1 916.386.5104 &nbsp;|&nbsp; Mobile: +1 916.317.5597"
)

# ── Mode ─────────────────────────────────────────────────────────────
# TEST MODE — same filters, but email goes to SENDER_EMAIL and Teams cards go to
# TEAMS_WEBHOOK_TEST. Cards are labelled as tests. Scope it with the filters below.
TEST_MODE            = False

# TEST_LIMIT — hard cap on how many emails a test run sends you, so an unfiltered
#   test cannot flood your inbox. 1 = one sample owner, 0 = no cap (send them all).
#   Combine with NOTIFY_ONLY to choose exactly whose email you want to see.
TEST_LIMIT           = 1

# ── FILTERS — scope who gets notified. Apply in ALL modes: dry run, test, real send.
#    All three stack: an owner must pass every active filter to be notified.
#    Leave a list empty to turn that filter off.

# NOTIFY_ONLY — whitelist. ONLY these CorpIDs get notified. [] = notify everyone.
#   Example: NOTIFY_ONLY = ["ABC1", "XYZ2"]
NOTIFY_ONLY          = []

# EXCLUDE_OWNERS — blacklist. These CorpIDs are ALWAYS skipped, even if whitelisted.
#   Example: EXCLUDE_OWNERS = ["ABC1"]
EXCLUDE_OWNERS       = []

# NOTIFY_DEVICES — only owners who have rules on these devices/sites.
#   Partial, case-insensitive match against the tab name, Device Name, and Policy Name.
#   A short site name matches all its variants, e.g. "Site_A" matches
#   "Site_A Sub", "Site_A_Station", etc.
#   An owner's rule list is also trimmed to just the matching devices, so they only
#   see rules for the site you are working on.
#   Example: NOTIFY_DEVICES = ["Site_Name_A", "Site_Name_B"]
NOTIFY_DEVICES       = []

# Tags to SKIP — base rules and standard exceptions don't need owner notifications
SKIP_TAGS            = {"BaseRule", "ToolsRule", "ToolsRules"}

# Shadowing: skip fully shadowed rules (already handled by decommission workstream)
SKIP_SHADOWED        = True
# ─────────────────────────────────────────────

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

DECISION_OPTIONS = "A) Recertify  |  B) Clean up / Remove  |  C) Review with team"


# ── Owner parsing ──────────────────────────────────────────────────────────────

def parse_owner_string(raw):
    owners = []
    if not raw or "Skip AMPS" in raw or "No APPID" in raw:
        return owners
    for line in raw.strip().split("\n"):
        line = line.strip()
        if not line or "=" not in line:
            continue
        try:
            ip_part, dict_part = line.split("=", 1)
            ip = ip_part.strip()
            data = ast.literal_eval(dict_part.strip())
            for app_id, hostname_map in data.items():
                for hostname, inner in hostname_map.items():
                    for app_key, fields in inner.items():
                        if not isinstance(fields, dict):
                            continue
                        for role, value in fields.items():
                            if not value or not isinstance(value, str):
                                continue
                            match = re.match(r"(.+?)\s*\(([A-Z0-9]+)\)", value)
                            if match:
                                name   = match.group(1).strip()
                                corpid = match.group(2).strip()
                                email  = f"{corpid}@pge.com"
                                owners.append({
                                    "email":    email,
                                    "name":     name,
                                    "corpid":   corpid,
                                    "role":     role,
                                    "app_id":   app_id,
                                    "ip":       ip,
                                    "hostname": hostname,
                                })
        except Exception:
            continue
    return owners


def extract_primary_owner(owners):
    priority = ["Client Owner", "Cyber Owner", "IT SME", "IT Lead", "IT SME backup"]
    for role in priority:
        for o in owners:
            if o["role"] == role:
                return o
    return owners[0] if owners else None


# ── Excel reading ──────────────────────────────────────────────────────────────

def detect_format(excel_path):
    """
    Decide which spreadsheet format we're looking at by inspecting the headers of the
    first sheet. Two supported shapes:
      "roles"  — new: one flat sheet with owner role columns (IT SME, IT Lead, ...)
                 and a Single APPID column. Owners are pre-resolved. (Sai's file.)
      "amps"   — old: many device-group tabs, owners parsed from the AMPS-owners
                 string columns. (The original 218-tab report.)
    """
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    first = wb[wb.sheetnames[0]]
    headers = set()
    for row in first.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = {str(h).strip() for h in row if h}
        break
    n_sheets = len(wb.sheetnames)
    wb.close()

    role_markers = {"IT SME", "IT Lead", "Client Owner"}
    if role_markers & headers and ("Single APPID" in headers or "App-ID" in headers):
        return "roles"
    if "Source AMPS owners list" in headers or "Destination AMPS owners list" in headers:
        return "amps"
    # Fall back on structure: one sheet with role-ish columns -> roles; many tabs -> amps
    return "roles" if (n_sheets == 1 and role_markers & headers) else "amps"


def _clean_appid(value):
    """Normalise an app id to APP-#### form. Returns '' if there's nothing usable."""
    if value is None:
        return ""
    s = str(value).strip()
    if not s or s.lower() in ("none", "no appid", "nan"):
        return ""
    s = s.upper()
    if s.startswith("APP-"):
        return s
    if s.isdigit():
        return f"APP-{s}"
    return s


def load_rules_roles(excel_path):
    """
    Reader for the new role-column format (one flat sheet). Keeps the app id and the
    resolved role owners on each rule so grouping can be by app id.
    """
    log.info(f"Loading spreadsheet (role-column format): {excel_path}")
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rules = []
    col = {}

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            col = {str(h).strip(): idx for idx, h in enumerate(row) if h}
            continue
        if not any(row):
            continue

        def g(field):
            idx = col.get(field)
            return row[idx] if idx is not None else None

        tags = str(g("Tags") or "")
        if any(t in tags for t in SKIP_TAGS):
            continue
        shadow = str(g("Shadowing Status") or "")
        if SKIP_SHADOWED and shadow == "FULLY_SHADOWED":
            continue

        app_id = _clean_appid(g("Single APPID") or g("App-ID"))

        # pull each configured role's "Name (CORPID)" cell as-is; parsed later
        roles = {}
        for role_col in ROLE_COLUMNS:
            val = g(role_col)
            if val:
                roles[role_col] = str(val).strip()

        rules.append({
            "sheet":        ws.title,
            "device_name":  g("Device Name"),
            "policy_name":  g("Application Acronym") or g("Application"),
            "rule_name":    g("Rule Name"),
            "source":       g("Source"),
            "destination":  g("Destination"),
            "service":      g("Service"),
            "application":  g("Application"),
            "action":       g("Action"),
            "tags":         tags,
            "last_hit":     g("Last Hit"),
            "shadowing":    shadow,
            "ticket_id":    g("Ticket ID"),
            "app_name":     g("Application Name"),
            "disabled":     g("Disabled"),
            "app_id":       app_id,
            "roles_raw":    roles,
            "business_justification": g("Policy Name + tag + FER close date + Business Justification"),
        })

    wb.close()
    log.info(f"Loaded {len(rules)} rules (role-column format)")
    return rules


_NAME_CORPID = re.compile(r"^(.*?)\s*\(([A-Za-z0-9]+)\)\s*$")


def _person_from_cell(cell, role):
    """Turn a 'Last, First (CORPID)' cell into an owner dict, or None."""
    if not cell:
        return None
    m = _NAME_CORPID.match(str(cell).strip())
    if not m:
        return None
    name, corpid = m.group(1).strip(), m.group(2).strip().upper()
    return {"name": name, "corpid": corpid, "email": f"{corpid}@pge.com", "role": role}


def build_appid_map(rules):
    """
    Group rules by app id (new format). For each app id, collect the set of people to
    notify from the configured NOTIFY_ROLES columns, de-duplicated by CorpID.
    Returns {app_id: {"recipients": [owner,...], "rules": [...], "app_name": str}}.
    """
    groups = {}
    for rule in rules:
        app_id = rule.get("app_id") or "UNKNOWN"
        grp = groups.setdefault(app_id, {"recipients": {}, "rules": [], "app_name": ""})
        grp["rules"].append(rule)
        if not grp["app_name"]:
            grp["app_name"] = rule.get("policy_name") or ""

        for role_col, cell in (rule.get("roles_raw") or {}).items():
            if role_col not in NOTIFY_ROLES:
                continue
            person = _person_from_cell(cell, role_col)
            if person and person["corpid"] not in grp["recipients"]:
                grp["recipients"][person["corpid"]] = person

    # flatten recipients dict -> list
    out = {}
    for app_id, grp in groups.items():
        out[app_id] = {
            "recipients": list(grp["recipients"].values()),
            "rules": grp["rules"],
            "app_name": grp["app_name"],
        }
    return out


def build_notification_units(rules):
    """
    Normalise the new role-column format into the {key: {...}} structure the send loop
    uses. Each unit is one email. Shape mirrors the old owner_map but carries a list of
    recipients (an app ID email goes to several role-holders) and a primary owner for
    logging/greeting.

    GROUP_BY = "app_id" -> one unit per app id, recipients = its NOTIFY_ROLES holders.
    GROUP_BY = "owner"  -> one unit per person, rules = everything they hold.
    """
    if GROUP_BY == "owner":
        # collapse to per-person, like the old map but sourced from role columns
        people = {}
        for rule in rules:
            for role_col, cell in (rule.get("roles_raw") or {}).items():
                if role_col not in NOTIFY_ROLES:
                    continue
                person = _person_from_cell(cell, role_col)
                if not person:
                    continue
                unit = people.setdefault(person["corpid"], {
                    "owner": person, "recipients": [person], "rules": [], "app_ids": set()
                })
                unit["rules"].append(rule)
                if rule.get("app_id"):
                    unit["app_ids"].add(rule["app_id"])
        return {cid: {"owner": u["owner"], "recipients": u["recipients"],
                      "rules": u["rules"], "app_ids": sorted(u["app_ids"])}
                for cid, u in people.items()}

    # default: group by app id
    appid_map = build_appid_map(rules)
    units = {}
    for app_id, grp in appid_map.items():
        recipients = grp["recipients"]
        if not recipients:
            log.warning(f"{app_id}: no recipients in the configured roles — skipping "
                        f"({len(grp['rules'])} rules). Check the role columns.")
            continue
        # primary owner = first recipient by role priority, for greeting/logging
        primary = recipients[0]
        units[app_id] = {
            "owner":      primary,
            "recipients": recipients,
            "rules":      grp["rules"],
            "app_ids":    [app_id],
            "app_name":   grp["app_name"],
        }
    return units


def load_rules(excel_path):
    log.info(f"Loading spreadsheet: {excel_path}")
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    rules = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = None

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = list(row)
                col = {h: idx for idx, h in enumerate(headers) if h}
                continue
            if not any(row):
                continue

            def g(field):
                idx = col.get(field)
                return row[idx] if idx is not None else None

            tags = str(g("Tags") or "")
            if any(t in tags for t in SKIP_TAGS):
                continue

            shadow = str(g("Shadowing Status") or "")
            if SKIP_SHADOWED and shadow == "FULLY_SHADOWED":
                continue

            rules.append({
                "sheet":        sheet_name,
                "device_name":  g("Device Name"),
                "policy_name":  g("Policy Name"),
                "rule_name":    g("ID on Device"),
                "source":       g("Source"),
                "destination":  g("Destination"),
                "service":      g("Service"),
                "application":  g("Application"),
                "action":       g("Action"),
                "tags":         tags,
                "last_hit":     g("Last Hit"),
                "shadowing":    shadow,
                "ticket_id":    g("Ticket ID"),
                "app_name":     g("Application Name"),
                "disabled":     g("Disabled"),
                "src_owners_raw": g("Source AMPS owners list"),
                "dst_owners_raw": g("Destination AMPS owners list"),
                "business_justification": g("Policy Name + tag + FER close date + Business Justification"),
            })

    total_sheets = len(wb.sheetnames)
    wb.close()
    log.info(f"Loaded {len(rules)} rules across {total_sheets} devices")
    return rules


# ── Owner aggregation ──────────────────────────────────────────────────────────

def build_owner_rule_map(rules):
    owner_map = {}
    for rule in rules:
        all_owners = []
        all_owners += parse_owner_string(rule.get("src_owners_raw", "") or "")
        all_owners += parse_owner_string(rule.get("dst_owners_raw", "") or "")
        owner = extract_primary_owner(all_owners)
        if not owner:
            continue
        email = owner["email"]
        if email not in owner_map:
            owner_map[email] = {"owner": owner, "rules": []}
        owner_map[email]["rules"].append(rule)
    return owner_map


# ── Filters ────────────────────────────────────────────────────────────────────

def rule_matches_devices(rule, device_filters):
    """True if a rule belongs to any of the requested devices/sites (partial, case-insensitive)."""
    haystack = " ".join(str(rule.get(f) or "") for f in ("sheet", "device_name", "policy_name")).lower()
    return any(d.strip().lower() in haystack for d in device_filters if d.strip())


def apply_filters(owner_map):
    """
    Narrow the owner map down using NOTIFY_ONLY / EXCLUDE_OWNERS / NOTIFY_DEVICES.
    An owner must pass every active filter. Returns a new map — the original is untouched.
    """
    whitelist = {c.strip().upper() for c in NOTIFY_ONLY if c.strip()}
    blacklist = {c.strip().upper() for c in EXCLUDE_OWNERS if c.strip()}
    devices   = [d for d in NOTIFY_DEVICES if d.strip()]

    if not (whitelist or blacklist or devices):
        return owner_map

    filtered = {}
    for key, data in owner_map.items():
        owner  = data["owner"]
        rules  = data["rules"]
        # a unit may notify several people (app-ID grouping) — filter on ALL of them
        recipients = data.get("recipients") or [owner]
        recip_ids  = {(r.get("corpid") or "").strip().upper() for r in recipients}

        # 1. Whitelist — unit passes if ANY of its recipients is whitelisted
        if whitelist and not (recip_ids & whitelist):
            continue

        # 2. Blacklist — drop any recipient that's blacklisted; skip the unit if none remain.
        #    (On the old format there's one recipient, so this reduces to the old behaviour.)
        if blacklist:
            recipients = [r for r in recipients
                          if (r.get("corpid") or "").strip().upper() not in blacklist]
            if not recipients:
                continue

        # 3. Device filter — unit must have at least one rule on a matching device,
        #    and its rule list is trimmed to only those rules
        if devices:
            rules = [r for r in rules if rule_matches_devices(r, devices)]
            if not rules:
                continue

        new_entry = dict(data)
        new_entry["owner"] = recipients[0] if recipients else owner
        new_entry["recipients"] = recipients
        new_entry["rules"] = rules
        filtered[key] = new_entry

    return filtered


def log_active_filters():
    """Print which filters are on so there is no doubt about who is in scope."""
    active = []
    if NOTIFY_ONLY:
        active.append(f"NOTIFY_ONLY: {NOTIFY_ONLY}")
    if EXCLUDE_OWNERS:
        active.append(f"EXCLUDE_OWNERS: {EXCLUDE_OWNERS}")
    if NOTIFY_DEVICES:
        active.append(f"NOTIFY_DEVICES: {NOTIFY_DEVICES}")

    if not active:
        log.info("  No filters active — all owners in scope")
        return
    for line in active:
        log.info(f"  {line}")


# ── Ticket ID cleaning ─────────────────────────────────────────────────────────

def clean_ticket_id(raw_ticket):
    ticket_lines = []
    for t in str(raw_ticket or "").strip().split("\n"):
        t = t.strip()
        if t and t not in ("0", "0.0"):
            try:
                ticket_lines.append(str(int(float(t))) if "." in t else t)
            except ValueError:
                ticket_lines.append(t)
    return ", ".join(ticket_lines) if ticket_lines else "N/A"


def truncate(text, length=60):
    text = str(text or "")
    return (text[:length] + "...") if len(text) > length else text


# ── Excel attachment builder ───────────────────────────────────────────────────

def build_excel_attachment(owner, rules):
    """Build an Excel file in memory with a Decision dropdown column for the owner to fill in."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Rules for Review"

    blue_fill  = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    light_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # Row 1 — instruction banner
    ws.merge_cells("A1:G1")
    ws["A1"] = "Firewall Rule Recertification — Please fill in the Decision column for each rule using the dropdown, then reply with this completed file."
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=11)
    ws["A1"].fill = blue_fill
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32

    # Row 2 — A/B/C legend
    ws.merge_cells("A2:G2")
    ws["A2"] = (
        "A) Recertify — Rule is still needed and should remain active     |     "
        "B) Clean up / Remove — Rule is no longer needed and can be disabled/deleted     |     "
        "C) Review with team — Unsure, would like to discuss with the NPS Automation team"
    )
    ws["A2"].font = Font(bold=True, color="003366", size=10)
    ws["A2"].fill = light_fill
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 28

    # Row 3 — column headers
    headers = ["Device", "Rule Name", "Source", "Destination", "Last Hit", "Ticket ID", "Decision (A / B / C)"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = blue_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="left")

    # Data rows starting at row 4
    for row_idx, r in enumerate(rules, 4):
        ws.cell(row=row_idx, column=1, value=str(r.get("device_name") or ""))
        ws.cell(row=row_idx, column=2, value=str(r.get("rule_name") or ""))
        ws.cell(row=row_idx, column=3, value=str(r.get("source") or ""))
        ws.cell(row=row_idx, column=4, value=str(r.get("destination") or ""))
        ws.cell(row=row_idx, column=5, value=str(r.get("last_hit") or "Never"))
        ws.cell(row=row_idx, column=6, value=clean_ticket_id(r.get("ticket_id")))
        cell = ws.cell(row=row_idx, column=7, value="")
        cell.alignment = Alignment(horizontal="center")

    # Dropdown on Decision column — rows 4 to end of data
    last_data_row = len(rules) + 3
    dv = DataValidation(
        type="list",
        formula1='"A) Recertify,B) Clean up / Remove,C) Review with team"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.sqref = f"G4:G{last_data_row}"
    ws.add_data_validation(dv)

    # Column widths
    col_widths = [25, 30, 40, 40, 15, 15, 28]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


# ── Email formatting ───────────────────────────────────────────────────────────

def get_first_name(owner):
    parts = owner["name"].split(",")
    return parts[1].strip().split()[0] if len(parts) > 1 else parts[0].strip()


def _recert_background():
    return """
    <p><strong>Background</strong></p>
    <p>A brief set of points signals the drivers for this effort:</p>
    <ol>
      <li>Heightened geopolitical strain and regional volatility are increasing the
      likelihood of opportunistic and state-aligned cyber activity, prompting a prudent
      reassessment of perimeter and inter-node controls.</li>
      <li>An evolving threat narrative — where automation, AI-assisted actors, and
      asymmetric techniques compress attack timelines — requires sharper situational
      awareness across all exposed attack surfaces.</li>
      <li>Growth in connectivity and exception-based access over time necessitates
      periodic rationalization to ensure firewall rules remain intentionally aligned to
      today's risk posture, not yesterday's needs.</li>
      <li>Proactive governance and operational readiness demonstrates disciplined control
      hygiene and reinforces that we are ahead of emerging risk vectors, not reacting to them.</li>
      <li>Compliance with PG&E and industry standards.</li>
    </ol>"""


def _recert_howto():
    return """
    <p><strong>How to Recertify your Applications</strong></p>
    <ol>
      <li>Open the SharePoint list linked above.</li>
      <li>Click the <strong>Edit in grid view</strong> button at the top of the page.</li>
      <li>Filter the list to show only your applications.
        <ul>
          <li>Filter by selecting a column heading and choosing "Filter by" (you can apply
          this to multiple headings).</li>
          <li>If you know your Application ID(s), that column will be the most effective.</li>
          <li>Alternatively, filter by your IT Director, then filter the IT SME, IT SME Backup,
          or Client Owner fields for your name. Filtering by IT Director first reduces the
          number of options in the other fields.</li>
        </ul>
      </li>
      <li>Review each rule and select the appropriate disposition in the
      <strong>Recertification</strong> column. You will need to <strong>double-click</strong>
      the cell in the Recertification column to bring up the pick list.</li>
    </ol>
    <p><strong>Note:</strong> The "Last Hit" column shows the last time a rule was used.
    If it is blank, the rule has never been used. If a rule has never been hit, or was last
    used more than 365 days ago, please consider it for deletion or mark it as Disaster
    Recovery if needed.</p>"""


def _recert_faq():
    return f"""
    <p><strong>Frequently Asked Questions</strong></p>
    <ol>
      <li>I can't access the recertification pick list, or I can't access the list at all.
        <ul><li>Make sure you click the "Edit in Grid View" button at the top left.</li></ul></li>
      <li>How do I get help if I don't know whether a rule is needed?
        <ul><li>Please come to office hours, or mark the line as "Need Assistance / Unsure"
        and NPS will reach out to you.</li></ul></li>
      <li>I don't think this rule should belong to me. How do I get assistance?
        <ul><li>Please come to office hours, or mark the line as "Need Assistance / Unsure"
        and NPS will reach out to you.</li></ul></li>
      <li>Can I update multiple rows at once?
        <ul><li>Yes. Update the first row with the option you want, select that cell and press
        CTRL-C to copy, then highlight the next cells and press CTRL-V to paste. You can paste
        into roughly 10-15 rows at a time.</li></ul></li>
    </ol>
    <p>The online FAQ will be updated as additional questions arise:
    <a href="{FAQ_LINK}">Firewall Recertification How-To and FAQ</a></p>"""


def format_email_recert(owner, rules, recipients=None):
    """
    The full recertification email (SharePoint mode) — matches the approved comms:
    the ask + four options, background, how-to, FAQ, office hours, signature.
    Static instructional content; owners record decisions in the linked SharePoint list.
    No attachment and no per-rule content in the body — the list holds the rules.
    """
    return f"""
    <html><body style='font-family:Arial,sans-serif;color:#333;line-height:1.5'>
    <p><strong>ASK</strong> — <strong>All</strong> firewall rules must be reviewed and
    recertified so that we have a full understanding of our attack surface area. Please respond
    to this request and update your assigned rules within <strong>{RECERT_WINDOW}</strong>. You
    may receive multiple emails for different groups of rules due to limitations of SharePoint
    lists. <strong>All</strong> rules <strong>must</strong> be recertified by
    <strong>{RECERT_DEADLINE}</strong>. If you <strong>do not</strong> respond by
    {RECERT_DEADLINE}, your rules may be <strong>disabled</strong>.</p>

    <p><a href="{SHAREPOINT_LINK}"><strong>Click here to open the recertification list</strong></a></p>

    <p>Review each rule and select the appropriate disposition in the Recertification column.
    Note that you will need to double-click the cell in the Recertification column to bring up
    the pick list:</p>
    <ol>
      <li><strong>Recertify</strong> — This rule will be kept and recertified.</li>
      <li><strong>Recertify — Disaster Recovery (DR)</strong> — This rule will be kept,
      recertified, and tagged as necessary for Disaster Recovery activities. This will prevent
      removal if it is unused for longer than 365 days.</li>
      <li><strong>Cleanup / Remove</strong> — This rule will be disabled and deleted via a
      scheduled change request.</li>
      <li><strong>Need Assistance</strong> — Network Protection Services (NPS) will reach out to
      you and provide assistance. Alternatively, you can join Office Hours for support.</li>
    </ol>

    {_recert_background()}
    {_recert_howto()}
    {_recert_faq()}

    <p><strong>OFFICE HOURS</strong></p>
    {OFFICE_HOURS_INFO}

    <hr>
    <p>{SENDER_SIGNATURE}</p>
    </body></html>
    """


def format_email_attachment(owner, rules, recipients=None):
    """
    Attachment-mode email (RESPONSE_METHOD = "attachment"): owners fill in the attached
    Excel and reply. Greets the group when app-ID grouped. SharePoint mode uses
    format_email_recert instead.
    """
    if recipients and len(recipients) > 1:
        firsts = [get_first_name(r) for r in recipients]
        if len(firsts) == 2:
            greeting = f"{firsts[0]} and {firsts[1]}"
        else:
            greeting = ", ".join(firsts[:-1]) + f", and {firsts[-1]}"
    else:
        greeting = get_first_name(owner)

    context = ""
    app_ids = sorted({r.get("app_id") for r in rules if r.get("app_id")})
    if recipients and len(recipients) > 1 and app_ids:
        who = "; ".join(f"{r['name']} ({r['role']})" for r in recipients)
        context = (f"<p>This request covers application <strong>{', '.join(app_ids)}</strong>. "
                   f"You are receiving it as a listed owner ({who}). "
                   f"Any one of you can complete the review.</p>")

    return f"""
    <html><body style='font-family:Arial,sans-serif;color:#333'>
    <p>Hi {greeting},</p>
    <p>As part of PG&E's annual firewall rule recertification process, the following
    firewall rules have been identified as requiring your review. These rules are
    associated with applications or assets you own.</p>
    {context}
    <p>Your full rule list (<strong>{len(rules)} rule{'s' if len(rules) != 1 else ''}</strong>)
    is attached as an Excel file.</p>
    <p><strong>Please open the attached file, fill in the Decision column for each rule,
    and reply to this email with the completed attachment:</strong></p>
    <ol>
      <li><strong>Recertify</strong> — The rule is still needed and should remain active.</li>
      <li><strong>Change / Clean up</strong> — The rule is no longer needed and can be disabled or removed.</li>
      <li><strong>Need assistance</strong> — You are unsure and would like to discuss with the NPS Automation team.</li>
    </ol>
    <p>If you have any questions or would like to schedule a review session,
    please contact the NPS Automation team directly.</p>
    <p>Thank you,<br><strong>NPS Automation Team</strong><br>PG&E Network &amp; Platform Security</p>
    <p style='font-size:11px;color:#888'>This is an automated notification from the NPS firewall recertification process.
    Rule count in this notification: {len(rules)}</p>
    </body></html>
    """


def format_email_reminder(owner, rules, days_outstanding=None):
    """Chase email. Assumes they have already read the original — no process explainer."""
    name   = get_first_name(owner)
    count  = len(rules)
    waited = (f" It has been <strong>{days_outstanding} days</strong> since that request."
              if days_outstanding else "")

    return f"""
    <html><body style='font-family:Arial,sans-serif;color:#333'>
    <p>Hi {name},</p>
    <p>We have not yet received your recertification decisions for
    <strong>{count} firewall rule{'s' if count != 1 else ''}</strong>
    associated with applications or assets you own.{waited}</p>
    <p>The same rule list is attached again so you do not need to look for the original
    email. <strong>Please fill in the Decision column for each rule and reply with the
    completed file:</strong></p>
    <ol>
      <li><strong>A) Recertify</strong> — The rule is still needed and should remain active.</li>
      <li><strong>B) Clean up / Remove</strong> — The rule is no longer needed and can be disabled/deleted.</li>
      <li><strong>C) Review with the team</strong> — You are unsure and would like to discuss with the NPS Automation team.</li>
    </ol>
    <p>If you have already replied, or if these rules belong to someone else now,
    let us know and we will update our records.</p>
    <p>If you would prefer to walk through the list together, contact the NPS Automation
    team and we will set up a review session.</p>
    <p>Thank you,<br><strong>NPS Automation Team</strong><br>PG&E Network &amp; Platform Security</p>
    <p style='font-size:11px;color:#888'>This is an automated reminder from the NPS firewall
    recertification process. Rule count in this notification: {count}</p>
    </body></html>
    """


# ── Sending ────────────────────────────────────────────────────────────────────

def send_email(to_email, owner, rules, dry_run=True, reminder=False, days_outstanding=None,
               recipients=None):
    """
    Send one recertification email. to_email may be a single address or a list — for
    app-ID grouping the same email goes to all the app's role-holders at once.
    SharePoint mode sends the full instructional email with no attachment (owners use
    the linked list); attachment mode sends the per-owner Excel to fill in and return.
    """
    to_list = [to_email] if isinstance(to_email, str) else list(to_email)
    to_header = ", ".join(to_list)
    rule_count = len(rules)
    plural     = "s" if rule_count != 1 else ""
    sharepoint = (RESPONSE_METHOD == "sharepoint")

    if reminder:
        subject   = f"Reminder: Firewall Rule Recertification Still Outstanding ({rule_count} rule{plural})"
        html_body = format_email_reminder(owner, rules, days_outstanding)
    elif sharepoint:
        subject   = "ACT: Execute Recertification of Firewall Rules"
        if EMAIL_SCOPE.strip():
            subject += f" - {EMAIL_SCOPE.strip()}"
        html_body = format_email_recert(owner, rules, recipients=recipients)
    else:
        subject   = f"Action Required: Firewall Rule Recertification ({rule_count} rule{plural})"
        html_body = format_email_attachment(owner, rules, recipients=recipients)

    if dry_run:
        who = to_header if len(to_list) <= 4 else f"{len(to_list)} recipients"
        log.info(f"[DRY RUN] Would email: {who} | {rule_count} rules | Subject: {subject}")
        return True

    msg = MIMEMultipart("mixed")
    msg["Subject"] = subject
    msg["From"]    = SENDER_EMAIL
    msg["To"]      = to_header
    msg.attach(MIMEText(html_body, "html"))

    # SharePoint mode carries no attachment — the rules live in the linked list.
    # Reminder and attachment modes attach the per-owner Excel.
    if not sharepoint:
        excel_bytes = build_excel_attachment(owner, rules)
        part = MIMEBase("application", "octet-stream")
        part.set_payload(excel_bytes)
        encoders.encode_base64(part)
        filename = f"Firewall_Rules_Recertification_{owner['corpid']}.xlsx"
        part.add_header("Content-Disposition", f"attachment; filename={filename}")
        msg.attach(part)

    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.sendmail(SENDER_EMAIL, to_list, msg.as_string())
        log.info(f"Email sent: {to_header} | {rule_count} rules")
        return True
    except Exception as e:
        log.error(f"Email failed to {to_header}: {e}")
        return False


# ── Response tracking ──────────────────────────────────────────────────────────

def load_responded_from_reporting_view():
    """
    Read the exported reporting view and return the set of CorpIDs who have finished
    responding — i.e. every rule they own has a non-blank Recertification decision.

    An owner with any blank Recertification cell still has outstanding work, so they are
    NOT counted as responded (they should still be emailed). Only owners whose rules are
    all decided are skipped.

    Returns a set of upper-case CorpIDs. Falls back to an empty set (skip nobody) if the
    file or the expected columns aren't present, so a misconfigured run never silently
    drops everyone.
    """
    path = Path(REPORTING_VIEW_PATH)
    if not path.exists():
        log.warning(f"SKIP_RESPONDED is on but no reporting view at {REPORTING_VIEW_PATH} "
                    f"— skipping nobody (everyone will be treated as outstanding)")
        return set()

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        log.error(f"Could not open reporting view {REPORTING_VIEW_PATH}: {e} — skipping nobody")
        return set()

    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    try:
        header = [str(h).strip() if h is not None else "" for h in next(rows)]
    except StopIteration:
        wb.close()
        return set()

    # locate the recertification column and the owner-role columns
    col = {h: i for i, h in enumerate(header)}
    recert_idx = None
    for h in col:
        if h.lower() == RECERT_COLUMN.lower():
            recert_idx = col[h]
            break
    if recert_idx is None:
        log.error(f"Reporting view has no '{RECERT_COLUMN}' column — skipping nobody")
        wb.close()
        return set()

    role_idxs = [col[h] for h in ROLE_COLUMNS if h in col]
    if not role_idxs:
        log.error("Reporting view has no owner-role columns — skipping nobody")
        wb.close()
        return set()

    # per owner: do they have any outstanding (blank) rule?
    has_outstanding = set()   # corpids with at least one blank decision
    seen = set()              # every corpid that appears
    for row in rows:
        if not any(row):
            continue
        decision = row[recert_idx] if recert_idx < len(row) else None
        decided = bool(decision) and str(decision).strip() != ""
        for idx in role_idxs:
            cell = row[idx] if idx < len(row) else None
            person = _person_from_cell(cell, "") if cell else None
            if not person:
                continue
            cid = person["corpid"]
            seen.add(cid)
            if not decided:
                has_outstanding.add(cid)

    wb.close()
    # responded = appeared in the view AND has no blank rules left
    responded = {cid for cid in seen if cid not in has_outstanding}
    log.info(f"Reporting view: {len(seen)} owner(s) seen, {len(responded)} fully responded "
             f"(will be skipped), {len(has_outstanding)} still outstanding")
    return responded


def load_responded():
    """
    CorpIDs that have already replied, so a send can skip them.

    When SKIP_RESPONDED is on, this is derived from the reporting-view export (the
    Recertification column). Otherwise it reads RESPONDED_FILE (a hand-maintained list),
    kept as a manual fallback.
    """
    if SKIP_RESPONDED:
        return load_responded_from_reporting_view()

    path = Path(RESPONDED_FILE)
    if not path.exists():
        log.info(f"No response file at {RESPONDED_FILE} — treating every owner as outstanding")
        return set()

    responded = set()
    with open(path, newline="", encoding="utf-8-sig") as f:
        for row in csv.reader(f):
            if not row:
                continue
            val = row[0].strip()
            if not val or val.lower() == "corpid":
                continue
            responded.add(val.upper())
    log.info(f"Loaded {len(responded)} responded owner(s) from {RESPONDED_FILE}")
    return responded


def load_notification_history():
    """First-notified timestamp per CorpID, read from previous notifications_*.csv logs."""
    history = {}
    # the running log, plus any older timestamped logs still lying around
    candidates = [Path(LOGS_DIR) / LOG_FILE]
    candidates += list(Path(LOGS_DIR).glob("notifications_*.csv"))
    candidates += list(Path(".").glob("notifications_*.csv"))   # logs from older runs
    for path in sorted(set(candidates)):
        if not path.exists():
            continue
        try:
            with open(path, newline="", encoding="utf-8-sig") as f:
                for row in csv.DictReader(f):
                    corpid = (row.get("corpid") or "").strip().upper()
                    stamp  = (row.get("timestamp") or "").strip()
                    if not corpid or not stamp:
                        continue
                    # rehearsals are not notifications — ignore dry runs and test sends
                    if str(row.get("email_sent")).lower() != "true":
                        continue
                    if str(row.get("dry_run", "")).lower() == "true":
                        continue
                    if str(row.get("test_mode", "")).lower() == "true":
                        continue
                    try:
                        when = datetime.fromisoformat(stamp)
                    except ValueError:
                        continue
                    if corpid not in history or when < history[corpid]:
                        history[corpid] = when
        except Exception as e:
            log.warning(f"Could not read {path.name}: {e}")

    if history:
        log.info(f"Read notification history for {len(history)} owner(s) from previous run logs")
    else:
        log.info("No previous notification logs found — no history to age reminders against")
    return history


def days_since(when):
    if not when:
        return None
    if when.tzinfo is None:
        when = when.replace(tzinfo=timezone.utc)
    return (datetime.now(timezone.utc) - when).days


def build_stats(owner_map, responded, history):
    """Roll-up used by the digest card."""
    outstanding = [(d["owner"], d["rules"]) for d in owner_map.values()
                   if d["owner"]["corpid"].upper() not in responded]
    ages = [days_since(history.get(o["corpid"].upper())) for o, _ in outstanding]
    ages = [a for a in ages if a is not None]

    holdouts = sorted(outstanding, key=lambda x: len(x[1]), reverse=True)[:5]
    return {
        "notified":          len(owner_map),
        "responded":         len(owner_map) - len(outstanding),
        "outstanding":       len(outstanding),
        "rules_outstanding": sum(len(r) for _, r in outstanding),
        "oldest_days":       max(ages) if ages else None,
        "holdouts":          [(o["corpid"], len(r), days_since(history.get(o["corpid"].upper())))
                              for o, r in holdouts],
    }


# ── Teams cards ────────────────────────────────────────────────────────────────

DECISION_HELP = ("✅ **A) Recertify** — Rule is still needed\n\n"
                 "🗑️ **B) Clean up / Remove** — Rule is no longer needed\n\n"
                 "💬 **C) Review with the team** — Need to discuss")


def _card(body, actions=None):
    """Wrap card body blocks in the Teams adaptive-card envelope."""
    content = {
        "$schema": "http://adaptivecards.io/schemas/adaptive-card.json",
        "type": "AdaptiveCard",
        "version": "1.4",
        "body": body,
    }
    if actions:
        content["actions"] = actions
    return {"type": "message",
            "attachments": [{"contentType": "application/vnd.microsoft.card.adaptive",
                             "content": content}]}


def _test_banner(owner=None):
    who = f" Generated for {owner['name']} ({owner['corpid']})." if owner else ""
    return {"type": "TextBlock",
            "text": f"🧪 **TEST — not an official notification.**{who}",
            "wrap": True, "color": "Warning", "weight": "Bolder"}


def _reply_action(owner, rule_count, reminder=False):
    """One-tap mailto reply. Avoids a spreadsheet round trip for simple answers."""
    subject = quote(f"Firewall Rule Recertification - {owner['corpid']} ({rule_count} rules)")
    return [{"type": "Action.OpenUrl", "title": "Reply by email",
             "url": f"mailto:{SENDER_EMAIL}?subject={subject}"}]


def build_owner_card(owner, rules, test_mode=False, reminder=False, days_outstanding=None):
    """Companion card (notify) or chase card (remind). Never carries the rule data itself."""
    name = get_first_name(owner)
    body = []
    if test_mode:
        body.append(_test_banner(owner))

    if reminder:
        waited = f" It has been {days_outstanding} days since that notice." if days_outstanding else ""
        body += [
            {"type": "TextBlock", "text": "⏰ Reminder: Firewall Rule Recertification Outstanding",
             "weight": "Bolder", "size": "Medium", "color": "Warning"},
            {"type": "TextBlock",
             "text": f"Hi **{name}**, we have not yet received your decisions for "
                     f"**{len(rules)} firewall rule{'s' if len(rules) != 1 else ''}**.{waited}",
             "wrap": True},
            {"type": "TextBlock",
             "text": "The original email has an Excel attachment with a Decision column. "
                     "Please complete it and reply, or let us know if you need a review session.",
             "wrap": True, "spacing": "Small"},
        ]
    else:
        body += [
            {"type": "TextBlock", "text": "🔒 Firewall Rule Recertification Required",
             "weight": "Bolder", "size": "Medium", "color": "Accent"},
            {"type": "TextBlock",
             "text": f"Hi **{name}**, you have **{len(rules)} firewall rule"
                     f"{'s' if len(rules) != 1 else ''}** awaiting recertification.",
             "wrap": True},
            {"type": "TextBlock",
             "text": f"📧 **Check your email.** A message from {SENDER_EMAIL} has an Excel "
                     f"attachment listing your rules, with a dropdown to record a decision "
                     f"for each one. This card is confirmation that the email is legitimate.",
             "wrap": True, "spacing": "Small"},
        ]

    body += [
        {"type": "TextBlock", "text": "**Decision Options:**", "weight": "Bolder", "spacing": "Medium"},
        {"type": "TextBlock", "text": DECISION_HELP, "wrap": True, "spacing": "Small"},
        {"type": "TextBlock",
         "text": "Reply to the email with the completed attachment, or contact the NPS Automation team.",
         "wrap": True, "spacing": "Medium", "isSubtle": True},
    ]
    return _card(body, _reply_action(owner, len(rules), reminder))


def build_announcement_card(owner_count, rule_count, test_mode=False):
    """Heads-up to the team channel before a notify wave goes out."""
    body = []
    if test_mode:
        body.append(_test_banner())
    body += [
        {"type": "TextBlock", "text": "📢 Firewall Rule Recertification — Notifications Going Out",
         "weight": "Bolder", "size": "Medium", "color": "Accent"},
        {"type": "TextBlock",
         "text": f"**{owner_count} rule owner{'s' if owner_count != 1 else ''}** are about to receive "
                 f"a recertification request covering **{rule_count} rules**.",
         "wrap": True},
        {"type": "TextBlock",
         "text": f"**What owners will see:** an email from {SENDER_EMAIL} with an Excel attachment "
                 f"listing their rules and a Decision dropdown (A / B / C).",
         "wrap": True, "spacing": "Small"},
        {"type": "TextBlock",
         "text": "**This is a legitimate internal request.** If someone asks whether it is phishing, "
                 "it is not — please point them here.",
         "wrap": True, "spacing": "Small", "weight": "Bolder"},
        {"type": "TextBlock", "text": DECISION_HELP, "wrap": True, "spacing": "Medium"},
        {"type": "TextBlock",
         "text": "Questions or escalations go to the NPS Automation team.",
         "wrap": True, "spacing": "Medium", "isSubtle": True},
    ]
    return _card(body)


def build_digest_card(stats, test_mode=False):
    """Status roll-up to the team channel. Contacts no owners."""
    body = []
    if test_mode:
        body.append(_test_banner())

    facts = [
        {"title": "Owners notified", "value": str(stats["notified"])},
        {"title": "Responded", "value": str(stats["responded"])},
        {"title": "Outstanding", "value": str(stats["outstanding"])},
        {"title": "Rules outstanding", "value": str(stats["rules_outstanding"])},
    ]
    if stats.get("oldest_days") is not None:
        facts.append({"title": "Oldest outstanding", "value": f"{stats['oldest_days']} days"})

    pct = f"{stats['responded'] * 100 // stats['notified']}%" if stats["notified"] else "n/a"
    body += [
        {"type": "TextBlock", "text": "📊 Firewall Rule Recertification — Status",
         "weight": "Bolder", "size": "Medium", "color": "Accent"},
        {"type": "TextBlock", "text": f"Response rate: **{pct}**", "wrap": True},
        {"type": "FactSet", "facts": facts, "spacing": "Medium"},
    ]

    if stats.get("holdouts"):
        lines = "\n\n".join(
            f"• **{c}** — {n} rules" + (f", {d} days" if d is not None else "")
            for c, n, d in stats["holdouts"]
        )
        body += [
            {"type": "TextBlock", "text": "**Largest outstanding owners:**",
             "weight": "Bolder", "spacing": "Medium"},
            {"type": "TextBlock", "text": lines, "wrap": True, "spacing": "Small"},
        ]

    body.append({"type": "TextBlock",
                 "text": f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')} by the NPS Automation team.",
                 "wrap": True, "spacing": "Medium", "isSubtle": True})
    return _card(body)


# ── Teams sending ──────────────────────────────────────────────────────────────

def teams_destination(test_mode=False, per_user=False):
    """Resolve which webhook a card should go to, and a label for logging."""
    if test_mode and TEAMS_WEBHOOK_TEST:
        return TEAMS_WEBHOOK_TEST, "test webhook"
    if per_user and TEAMS_USER_WEBHOOK:
        return TEAMS_USER_WEBHOOK, "per-user flow"
    return TEAMS_WEBHOOK, "channel webhook"


def post_teams(payload, webhook, label, description, dry_run=True, recipient=None):
    """POST one card. recipient is passed through for per-user Power Automate flows."""
    if dry_run:
        log.info(f"[DRY RUN] Would post Teams card ({label}): {description}")
        return True

    if not webhook or webhook.startswith("<"):
        log.warning(f"Teams webhook not configured ({label}), skipping: {description}")
        return False

    if recipient:
        payload = dict(payload, recipient=recipient)

    try:
        resp = requests.post(webhook, headers={"Content-Type": "application/json"},
                             data=json.dumps(payload), timeout=10)
        if resp.status_code in (200, 202):
            log.info(f"Teams card sent ({label}): {description}")
            return True
        log.error(f"Teams post failed ({label}) for {description}: {resp.status_code} {resp.text}")
        return False
    except Exception as e:
        log.error(f"Teams error ({label}) for {description}: {e}")
        return False


def send_teams(owner, rules, dry_run=True, test_mode=False,
               reminder=False, days_outstanding=None):
    """Send one owner-directed card, per-user if a flow is configured."""
    webhook, label = teams_destination(test_mode=test_mode, per_user=True)
    if test_mode and not TEAMS_WEBHOOK_TEST and not dry_run:
        log.warning("TEAMS_WEBHOOK_TEST is empty — test card is going to the real destination")
    if label == "channel webhook" and not test_mode:
        log.warning(f"No per-user flow configured — {owner['corpid']}'s card is going to the channel")

    payload = build_owner_card(owner, rules, test_mode=test_mode,
                               reminder=reminder, days_outstanding=days_outstanding)
    return post_teams(payload, webhook, label,
                      f"{owner['corpid']} | {len(rules)} rules"
                      + (" | reminder" if reminder else ""),
                      dry_run=dry_run, recipient=owner["email"])


def send_teams_channel(payload, description, dry_run=True, test_mode=False):
    """Send one channel-directed card (announce / digest)."""
    webhook, label = teams_destination(test_mode=test_mode, per_user=False)
    return post_teams(payload, webhook, label, description, dry_run=dry_run)


# ── Logging ────────────────────────────────────────────────────────────────────

def write_log(log_file, rows):
    Path(LOGS_DIR).mkdir(exist_ok=True)
    log_file = Path(LOGS_DIR) / log_file
    fieldnames = [
        "timestamp", "email", "name", "corpid", "role",
        "rule_count", "attachment_sent", "dry_run", "test_mode",
        "email_sent", "teams_sent", "device_list"
    ]
    # Append to one running log; write the header only when the file is new.
    new_file = not log_file.exists()
    with open(log_file, "a", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        if new_file:
            writer.writeheader()
        writer.writerows(rows)
    log.info(f"Logged {len(rows)} row(s) to: {log_file}")


def ensure_workspace(log_rows):
    """After a genuine send, create the folders and files used to track replies."""
    if DRY_RUN or TEST_MODE:
        return
    if not any(r["email_sent"] for r in log_rows):
        return

    if not Path(RESPONSES_DIR).exists():
        Path(RESPONSES_DIR).mkdir(exist_ok=True)
        log.info(f"Created {RESPONSES_DIR}/ — save replied-to spreadsheets here")

    Path(LOGS_DIR, "discarded").mkdir(parents=True, exist_ok=True)

    responded = Path(RESPONDED_FILE)
    if not responded.exists():
        responded.write_text("corpid\n", encoding="utf-8")
        log.info(f"Created {RESPONDED_FILE} — add one CorpID per line as replies arrive")


# ── Main ───────────────────────────────────────────────────────────────────────

def notify_run(owner_map, reminder=False, responded=None, history=None):
    """Contact owners: email + optional companion/reminder Teams card."""
    responded = responded or set()
    history   = history or {}
    log_rows  = []

    if TEST_MODE and TEST_LIMIT > 0 and len(owner_map) > TEST_LIMIT:
        log.warning(f"TEST MODE — capping at {TEST_LIMIT} of {len(owner_map)} owner(s). "
                    f"Raise TEST_LIMIT, or set it to 0, to send them all.")
        owner_map = dict(list(owner_map.items())[:TEST_LIMIT])

    for key, data in owner_map.items():
        owner, owner_rules = data["owner"], data["rules"]
        # recipients: the list for app-ID units, or just the single owner (old format)
        recipients = data.get("recipients") or [owner]
        recipient_emails = [r["email"] for r in recipients]
        days = days_since(history.get(owner["corpid"].upper()))

        if TEST_MODE:
            to_list = [SENDER_EMAIL]                # test send goes to you / the group
        else:
            to_list = recipient_emails
        email_ok = teams_ok = False

        if NOTIFY_EMAIL:
            email_ok = send_email(to_list, owner, owner_rules, dry_run=DRY_RUN,
                                  reminder=reminder, days_outstanding=days,
                                  recipients=recipients)
        if NOTIFY_TEAMS:
            teams_ok = send_teams(owner, owner_rules, dry_run=DRY_RUN, test_mode=TEST_MODE,
                                  reminder=reminder, days_outstanding=days)
            if not DRY_RUN:
                time.sleep(TEAMS_DELAY_SECONDS)

        devices = list({r["device_name"] for r in owner_rules if r["device_name"]})
        log_rows.append({
            "timestamp":       datetime.now(timezone.utc).isoformat(),
            "email":           "; ".join(recipient_emails),
            "name":            owner["name"],
            "corpid":          owner["corpid"],
            "role":            owner["role"],
            "rule_count":      len(owner_rules),
            "attachment_sent": True,
            "dry_run":         DRY_RUN,
            "test_mode":       TEST_MODE,
            "email_sent":      email_ok,
            "teams_sent":      teams_ok,
            "device_list":     "; ".join(devices[:5]) + ("..." if len(devices) > 5 else ""),
        })

    return log_rows


def _run_current_config():
    mode = RUN_MODE.strip().lower()
    if mode not in ("notify", "remind", "announce", "digest"):
        log.error(f'RUN_MODE "{RUN_MODE}" not recognised. '
                  f'Use one of: notify, remind, announce, digest')
        return

    log.info(f"RUN_MODE: {mode}")

    # If owners are being directed to SharePoint, make sure a real link is set —
    # otherwise every email would carry the placeholder. Warn loudly; don't send blind.
    if mode in ("notify", "remind") and NOTIFY_EMAIL and RESPONSE_METHOD == "sharepoint":
        if not SHAREPOINT_LINK or SHAREPOINT_LINK.startswith("<"):
            log.error("RESPONSE_METHOD is 'sharepoint' but SHAREPOINT_LINK is not set. "
                      "Set the list link for this network (ODN/GDN), or switch "
                      "RESPONSE_METHOD to 'attachment'.")
            return
        log.info(f"Response method: SharePoint list — {SHAREPOINT_LINK}")
    elif mode in ("notify", "remind") and NOTIFY_EMAIL:
        log.info("Response method: Excel attachment (reply with completed file)")

    if DRY_RUN:
        log.info("=" * 60)
        log.info("DRY RUN MODE — no notifications will actually be sent")
        log.info("Set DRY_RUN = False in CONFIG to send for real")
        log.info("=" * 60)
    if TEST_MODE and not DRY_RUN:
        log.info(f"TEST MODE — email goes to {SENDER_EMAIL}, Teams cards are labelled as tests")

    fmt = detect_format(EXCEL_PATH)
    log.info(f"Spreadsheet format detected: {fmt}")

    if fmt == "roles":
        rules = load_rules_roles(EXCEL_PATH)
        owner_map = build_notification_units(rules)
    else:
        rules = load_rules(EXCEL_PATH)
        owner_map = build_owner_rule_map(rules)
    total_owners = len(owner_map)

    owner_map = apply_filters(owner_map)
    log.info(f"Notification units before filters: {total_owners}")
    log.info(f"Notification units after filters:  {len(owner_map)}")
    log_active_filters()

    if not owner_map:
        log.warning("No owners matched the filters — nothing to do. Check your filter lists.")
        return

    # Responded set: always in remind/digest; in notify too when SKIP_RESPONDED is on
    # (so a repeat send only emails people who still have blank decisions).
    load_resp = mode in ("remind", "digest") or (mode == "notify" and SKIP_RESPONDED)
    responded = load_responded() if load_resp else set()
    history   = load_notification_history() if mode in ("remind", "digest") else {}
    log_rows  = []

    # ── Channel-only modes: no owner is contacted
    if mode == "announce":
        payload = build_announcement_card(len(owner_map), sum(len(d["rules"]) for d in owner_map.values()),
                                          test_mode=TEST_MODE)
        send_teams_channel(payload, f"announcement | {len(owner_map)} owners",
                           dry_run=DRY_RUN, test_mode=TEST_MODE)

    elif mode == "digest":
        stats   = build_stats(owner_map, responded, history)
        payload = build_digest_card(stats, test_mode=TEST_MODE)
        send_teams_channel(payload, f"digest | {stats['outstanding']} outstanding",
                           dry_run=DRY_RUN, test_mode=TEST_MODE)
        log.info(f"Responded: {stats['responded']} | Outstanding: {stats['outstanding']} "
                 f"| Rules outstanding: {stats['rules_outstanding']}")

    # ── Owner-directed modes
    elif mode == "remind":
        before = len(owner_map)
        owner_map = {e: d for e, d in owner_map.items()
                     if d["owner"]["corpid"].upper() not in responded}
        log.info(f"Skipped {before - len(owner_map)} owner(s) already recorded as responded")

        if REMIND_AFTER_DAYS > 0 and history:
            aged = {}
            for e, d in owner_map.items():
                days = days_since(history.get(d["owner"]["corpid"].upper()))
                if days is not None and days >= REMIND_AFTER_DAYS:
                    aged[e] = d
            log.info(f"Skipped {len(owner_map) - len(aged)} owner(s) notified less than "
                     f"{REMIND_AFTER_DAYS} days ago or with no recorded notification")
            owner_map = aged

        if not owner_map:
            log.warning("Nobody is due a reminder right now.")
            return
        log.info(f"Reminding {len(owner_map)} owner(s)")
        log_rows = notify_run(owner_map, reminder=True, responded=responded, history=history)

    else:  # notify
        if SKIP_RESPONDED and responded:
            before = len(owner_map)
            owner_map = {e: d for e, d in owner_map.items()
                         if d["owner"]["corpid"].upper() not in responded}
            log.info(f"Skipped {before - len(owner_map)} owner(s) who have already responded; "
                     f"emailing {len(owner_map)} still outstanding")
            if not owner_map:
                log.warning("Everyone in scope has already responded — nothing to send.")
                return
        log_rows = notify_run(owner_map)

    if log_rows:
        write_log(LOG_FILE, log_rows)
        ensure_workspace(log_rows)
        if not DRY_RUN and not TEST_MODE:
            log.info(f"To void a run, delete its rows from {LOGS_DIR}/{LOG_FILE} — "
                     f"those owners will stop counting as notified")

    print("\n" + "=" * 60)
    print("  RUN SUMMARY")
    print(f"  Run mode:                 {mode}")
    print(f"  Total rules loaded:       {len(rules)}")
    print(f"  Owners found:             {total_owners}")
    print(f"  Owners in scope:          {len(owner_map)}")
    if log_rows:
        print(f"  Owners contacted:         {len(log_rows)}")
        print(f"  Rules in notifications:   {sum(r['rule_count'] for r in log_rows)}")
    if mode in ("remind", "digest"):
        print(f"  Recorded as responded:    {len(responded)}")
    print(f"  NOTIFY_ONLY:              {NOTIFY_ONLY or 'all owners'}")
    print(f"  EXCLUDE_OWNERS:           {EXCLUDE_OWNERS or 'none'}")
    print(f"  NOTIFY_DEVICES:           {NOTIFY_DEVICES or 'all devices'}")
    print(f"  Email notifications:      {'enabled' if NOTIFY_EMAIL else 'disabled'}")
    print(f"  Teams notifications:      {'enabled' if NOTIFY_TEAMS else 'disabled'}")
    print(f"  Test mode:                {TEST_MODE}")
    print(f"  Dry run:                  {DRY_RUN}")
    if log_rows:
        print(f"  Log file:                 {LOGS_DIR}/{LOG_FILE}")
    print("=" * 60)


def _apply_batch(n):
    """Point the module-level config at batch n's spreadsheet, link, scope, reporting view."""
    global EXCEL_PATH, SHAREPOINT_LINK, EMAIL_SCOPE, REPORTING_VIEW_PATH
    b = BATCHES[n]
    EXCEL_PATH          = str(SCRIPT_DIR / b["excel"])
    SHAREPOINT_LINK     = b["link"]
    EMAIL_SCOPE         = b["scope"]
    REPORTING_VIEW_PATH = str(SCRIPT_DIR / b["reporting"])


def main():
    # Which batches to run: one, or all of them in sequence.
    if BATCH == ALL_BATCHES:
        order = sorted(BATCHES.keys())
        log.info(f"BATCH = {ALL_BATCHES} (ALL) — running {len(order)} batches in sequence: "
                 + ", ".join(BATCHES[i]["scope"] for i in order))
    elif BATCH in BATCHES:
        order = [BATCH]
    else:
        log.error(f"BATCH is set to {BATCH}, which is not a valid choice. "
                  f"Use 1-{max(BATCHES)} for a single batch, or {ALL_BATCHES} for all.")
        return

    for i in order:
        b = BATCHES[i]
        log.info("=" * 60)
        log.info(f"BATCH {i}: {b['scope']}  ({b['excel']})")
        log.info("=" * 60)
        # guard: don't run a batch whose link was never filled in
        if not b["link"] or b["link"].startswith("<"):
            log.error(f"Batch {i} ({b['scope']}) has no SharePoint link set — skipping it.")
            continue
        _apply_batch(i)
        try:
            _run_current_config()
        except Exception as e:
            log.error(f"Batch {i} ({b['scope']}) failed: {e}")
            if BATCH == ALL_BATCHES:
                log.info("Continuing to the next batch...")
                continue
            raise


if __name__ == "__main__":
    main()
