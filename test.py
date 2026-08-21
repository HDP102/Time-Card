#!/usr/bin/env python3
"""
FireMon Device Health Report -> Excel

Self-contained. The Device Health Report of 21 Aug 2026 (143 devices) is embedded
below, so running this with no arguments reproduces the workbook exactly:

    python firemon_health_to_excel.py

To use a newer export instead, pass it as an argument:

    python firemon_health_to_excel.py DeviceHealthReport.txt [output.xlsx]

Export the source from Security Manager:
    Reports -> Reports Library -> Health Check -> Device Health Report
    Run against All Devices, then export/copy as text.

Requires openpyxl (pip install openpyxl).
"""

import base64
import gzip
import re
import sys
from collections import Counter
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

FONT = "Arial"

# ---------------------------------------------------------------- embedded report
# Device Health Report, 21 Aug 2026 8:13:14 PM UTC, 143 devices.
# gzip + base64 of the raw text export. Used when no input file is given.

_EMBEDDED = """
H4sIAEK4iGoC/+1daXPbSJL9TP0KROxGyJ4V6DqBAmOmo2lJdmtH14iye7ZnOhw4ChK3KUALkLLl
6B+/WbgIXuJt8UBYlkgcdWa+fJl1ncintiu1X6Td6d5rN/IxjLoHzd5dL+5qBB9pBBHjSBMNTBuY
adcX2qfb44PrKHxqezLW7EALn2T01JZftdDXuvdSi7t2V6ovXpJyrIWR9mAH9p18kEE3ud0Og/gI
3vU0FxJoB26n50ntPi2Cey/dP7RIxr1ON64ffGh3ujJqaJdhIA9O0iShJDT7XDuRsRu1H1WateMO
lFpGtYt+dmfXWtPzILW49lkGXhjV0poeNL/3IqkfX326vL35H+1DO7ADt213tJN2bKtEoClcVbVn
7c3ZSUPD6G0NbiWXpKc5z1qSQK12GWoD2dUu2m4UxqHfrR0cR+1u27U7B1nrHidVu0mrdvDx9PL0
pnl+cHL6+ez4VDuHX5et05ODk2Pt483VJyh4q3X28RKuNDXP7tqaG3Y60u1Ce95FYe9Ru7djzZEy
0Ow4bt8FUKxuCF3QjrOm1958Rrh1ety8PkeGWXfDXtCNnjt2AB8f9I8qjbcHN6e3N2enn6Ec583W
rVb6Cs2g9R4hZ0g4DDTjHbHeEUOzuyANjDaokoa61uxBpweqmqoLNN9ud6TX0JrNk9Zty0QIEcwb
2lnwZHfa0N+ddiIE0o1kV3tMxcira6dBDK2Zyk96z5Ht4A6+wNPtILkRyf/rSSgT1E99HUwKku/J
Iy0Iu2Nunp0caT40m51fsD0vbS378VE7tG3D8QzKdOpZls58InSLuoaOOVxwbJMICx/WtdvIhjZV
wsAptQWRVPdsD+vMNDzdZhbVOfF96Xg+dRDSjsMokp20VdRLNrNN7glH9x0udIY9ogtTQFaIcI+7
nFqOod22H6CK9sNjI1E8HRk6seBj0uDkt7yPPp+1zq4uD26hph3VTRH0d5w0P9SyLAH3tge1hipH
bQlNpMkoCqODm/zxRLINgoyD2+dH2VCdGT5AkV0Q0vwVXbsKQHLt4A70D4ThnSqjenQIIniDWg0o
bAYRn2KltSfHX+w0TekdpIlAw4AcNjR0kCpCo5TXaVK841+alx9PtZPT29PjW1XP7MLF1eXZ7dXN
2eXHg9PAdjoqyezR5m0zSz7WvoKGQv27oCqppgLG2EXFirY4+NRqfjw9OL/6WE74EgQIhKID8gwZ
pM+kyQ/dOQ/v7iD54zAIICOFaNqbRGUI0+7DXhS/VVWchDKAZhqAqJeCC6+wZQhb6N5iCyWGID62
delTojMipO4gDlghmGdw36fWGGxhlie48BxdmqbUmQe5CIl83bWwjTBCpkTWa2ELXxm2mBW2zIIt
Cld6yWspuhgVupTRhe01c7GIZNi0FAlBOkO+pVsYcEMw4Qsu+VjmYmBfehZxdeFgBuiCiO64DtYR
xy7DwgJYeom5sHWiC14Zuhhbji77KM42I9Ql2NN97oOxNDjXHS4dnVoeNz1pEG+csaSEIBNyccDW
AnunGIg4BnF2kGXbFqY2RqPiLHSClTgjo8HJb3VtGGTEO4ILkFEPKQpTxvhPMUBtCocK/lQDJRAY
SVe2nyAR1TAEHngGQbhvu/ea/OZK6WWNHAZ++66nELx7D+7sfdjxVH/EstuF/olXZzNa97bKpaW8
euW+t3pO4WGnBsWsDErJoBC2z3SVM8vlBua6KU1XZ4I5uuVZCFxhBzumbSN7jAa6yMSEcEsnwgGD
wsEAOQawXduktmn5DhGmOdmgkLUaFFIZlB9LV//RTFFFVKhSRpW9doJtgqhPiKVTzH1gnMIEbLA4
QIXPkOO6RI5BFdNBDraFr1MXKVSxDN1mDLJysStsKrgr8AuoslYnmFWo8mNRpWUHnhN+y4L3uMKW
AlvMBsL77AJTG/iJazNdEEMAYzEMXfiuq0vf9cH7FdQcgy0C+a4wiK0bgnvgaFhSFxbAh+05BpGO
69kUTcAWbCUNTteGLeaqsIWSClvmwRZdu5Tfuh9l5hZhRiuYKVEYSvaZwmBsUwm+kc5MYQNi+FR3
bOTqnmRARSyP8nGhCdNiro1t3fZ9oDAEY90SwtMN1weMsh0kQUAnUhjV4OujMGJlMMOqSNvWibOk
psEwwrrnK5ddEKQLybBuSOTYXDA2LnDsGcKwGaTtCniU+RLrwnUFcHlMLNOxJXemRNrwLJE2vImR
tvfnF/rJRUtP/p63wIycX73XO+EdyDzYh2zgl72tQXJJsA2uQUcllVTlO75oglphVMdGHSOzTgiq
1cpfeS01FddhOwA7kpqP5600I8PWeQDAhm7OoPgT3hjDA34EixgWhJObESmgc0sBQfCf1Q1aCcH2
CcHFWCEwFhUCLioh2CohGO55vrgRQFXXb0XXQ1OGtv4Uf9P9rzgPSplva1fRIzDSlIKOdDa2auX3
VMcLuAMPsMFu/xRAaWQQA9kdyOiLC132h617ke6EkScjuJjNCSFT8gbJGvNyDdWTf3Nn7z187+fN
58k7fXOpjOOveYujeXOOF89ZfusWVbaMeTJO31wq47zKwpw34yVq3I7DosbmXBmnb86bcfJtSLDN
WVp68MWFsi0LNEZ81kwXk+bi1bxbrbkyjBfLsCzAppg1w8WEt3g1r6GJ5slwwRp+tYO+iuJZM0zf
mitDkgOQsTDk89kyegHyrfVDPpkA+RjRNWM+mYD5llgz5JMJkC/YmiGfTIB8kJY1Yz6ZgPkYiTWD
PhkP+pD2elGfjEN9TNaI+mQM6htojahPxqC+IGtEfTIO9fEaUZ+MQX1j1ajfsZ2C4ZOpiItryfMK
5k1ax3U85Mpv9eBQq+e6Mo79XqfznIfUIYs8vpqGid9c2EHP7rzVPrQ7KvaahP3FkXYbdu2OFre/
S5AKVrcu3o8Ggsk7zN4Rng5qY5wFgqcNfwRh9GB3IGHvSHN6KrYe90ZGBNoBZCCTpYthJPOljT60
X/hVBeltcNLjhhbI7tcw+kMLnf+F9+IjLU5nTfYv9OBK/1vmNyY1Ly4+taMuNIEGzdmVEVyArpeR
b2drJ5PL8eB4DGP5aMx1CAk+a2dBDO3VKQ++fJBO1LOjZw2TZPiFH2XhcqY1y8MvoWsZjJGflVxa
BIQShJOyoYEYyyhGYi6L9lvtIIzXi1TLPqYVaqcVSppqkoeeBPQDEIl+vD8P8Y8VFp4JCzQDbSAT
mmHjRg0SOMjYIkVzw4c1CB9ngQ1ZPUmV7FDwgaDpTLT/VpGDYVQANQpQHM8CT9jYF3jCFp8dnniO
TgrBWQMNjQ63PYtyIVJ4qpM6JRPHiHcCmfDmIlPhPavPQxR8Kp4QWht4bTyTGodYSXZllkjwbHm9
xBH3ALha7r30eiDiL2MXnUiuBu0lVlI5HJUvcCsGmWtDCkOTSxZVxBlSnqR6u8UGhgJIU0V/gtk2
p6oZGafVU4coZtPqiiVkmsan84RsY409cWPowm4MoYNuzFSeAJxkA4nCTjIEMmqyyfImuwCtACH/
a8d+7CCUuTPmDMFWp/OQQyKpG2wk1jIu+Qx2TbJA8lZFOFLFw2gQBzFG42CQvsOiD4Osgc1RwrGj
7hIyjKk4eGFHoLa4j4HgLiE6EsoxsYn6oRxex0wM4SDe+lAOSIpVSApiatOhDYTDAGH77i7qRgVK
YRUQLubv1AEg1GKB3V0CYDQQaRAEnVNPZvmC6GcNrXr1r4/y2yOUs/7Yff4SP9pfg3ryO1MllQD6
ZvpC+sSXgljoJ6jZw4MdeA3tXS+O3jnt4F0c34P23oHy/uuwfPHwSDvUH9VvQtTv71/8diQfwuDn
ovUPfwfA8H01CyuBFIyQ5t7bSe85h4ea7Sf7zv3VhbuxdpiXVlUgWQsf12/PLk6vPt0e/qQ92F33
Pt2jLv38pR148lt2BfyXrtr2rhdnF/yOffdFhj6wG7sTS+2x7Sm9pJhiC8rQ7nhffLgCnzthrKaW
p4910zZUkUHouk77oT2thKdXH6B0nfDOBzzOMs++fQHM84YuxTLwilp8Sx8gCNoFxAaA07VjmZcl
lgqPAD298GsK68lrUCz72ZGgXzJNDKw4T68m7ZlUSF3FpYvw89AOQIDSG2nSqmb5Jygr3GqA/iqR
fISivnEOoc6A1l4yRx76NsP/QJUilodvNTz0PCBkYAOS/gkfLtUHeIYMPfOv68ffs3Qbf16DYh1f
nZymn369ujlp/HncsZ+SN+nwm//xl//6/T/hBhu68dO//x3/BW6sZ1kGFvnK0sL49pdKvLgUI1VO
a9CArGspxmZNxrtE+MOv5wru+shMhpdulcib1jpuaeNmYwLHs3vQBOXpGsm90Sl6e0D/ioVBw/Rv
1IIPLEelpIH5LFxvvIbkDDAtxTgtIRTn/Cp108eryH+re+BoErO/XIlA2cigjmSGZJhVITZCq9IS
rW69Ur/l1YIltalqQjQLdVRECzBb8dt5V02mzEiRtpSBtQNfVUJVoJ7JamrB+hu7av/63Prn4Hzm
36cs1lERfrGJi3XKiPAlU+nhiYV8EYQAx3A4uZ0OkeWcPxOWQvJehAAMSgZEXswS9loUAygS1gwY
0I82iT4IILNB8OColBo0FxYeAgGxdxAwsqRhpyCgPN3Losvp/0tzwCrlX7fyW2IZ5aeV8u+n8vfn
1BrLa//UuPa4UpSnYxpsuUK8NEmzgqC18w++DAShCoL2D4LKI2jcWl7540r5X0v5jUr5K+WfS/nL
yxQwWlL7X1q9MI5/FPON6KrjonyGzCeEYDBia43BvFiWwaU/eH3O4NRSFGt7rfVx0soirNkigC4s
YxFYZRH2xiKQcb6gQOvzBV9EoIHlenR9nPTFQgws3EbrM40VDK49MEeXgUFSweCewCArhuqxWI6S
duPuGFqKq+H6DRquN5cYrqezDdfTarh+J1DhS0mth/xFtpi7iGvjkpyVpQwVqew2Urx8eaphxEnb
TXOuVs6sE5fITGRlHC4Bl0LGLLgkrC2eRjTb7O3thphyAIaj1ehzvCi4lB1CxpcvzHxO4ZjC5C3D
jNUUZuGWKTuInCxfmPnip6yIn2JrHWSVzFCAF+wiNdZuFytTtG5TZCxhiqxZTBElVmWKNswUkYnU
kq6XWk6FmLJdZGu2ixW4rBdcaHGq7gLgMpv/jSpo2WRoKRNLaq2XWFbavG5tRuvWZkorr3Xj9Tm3
zpSs1zer9Hnd1J8sQf15ZZ23XpvLwQ2G1xvcqLR53dZ5iZgyZrNoMxmzeWulz6+zI0j7MR7YEUSF
4vo7gmC0H1uC0OW2BJHM9z1DILHiLUEw2sg9QeiMe4Ksa4cJY5kdJijZxx0mQNMfe86gptOyppu7
ruiigVC+909xEioAqDo3FQS3viZhtcwFhTUpL0V7Lqz5sBXlZWEVe2CVUG6VfqCwCro4siJ1tvae
CutTdxBZzbKw8opCzUChTJN50nFXvqsa30gGxfDrMiixFIPadz0vjJIo67mxH3rOl9NzjznCxmhJ
Pf9tRM+NjdRzzl5Xz81KzxfQ86fHoLyXM6YDu6QSvOuKbiaKjpdTdORxITjx5IoNOsEbqOmMMPKq
mm4tGhMx93fXzZKm5xZdTRQqMXezMumzaDpyXEP62Fw1dTc306bz17Xp1lI2ne+nppPhnc9prumm
UUd1ulcBpUt4+Ozky/HN6cnp5e1Z87y1LllFS8WT9jT4Sbz24Jgc4X1ZJfV9oJ8op5+t1i/a3+Wz
Qn3obq9/Aaopu8mwqa0OzUgTKZWurl2EXtt/ToZLswLno6OqIjIZQE6S+0OqRalwc13cDC3OzfZX
C/gIYrOcm5mijolZx9aewPbuqIJpVQZhAVUYOlyKz3C4fOvm6stxBxpRRmWdIbw6drI6BWrGU6AM
U8x6ChTLj4EyG0SkHuW8J3pjvHHHQOUpb968LT50IBwXS0ACnnjunDVihHHZbWL7ZH9/oNuEl7KS
+MdYyXmnYMZFZVay48gqLSzGIy5XPg3SZPvgcg3I+Q+cb0K2QdA3jA5iVoDykATkGE1yR0mw1FHK
xTeV2uc9O/V4mOgRyv4+9tRjszjSFlvqePEF9pnr8z4tLhV4UPKpMdfO62bpiE0o19AWc/YfsV1/
APWwffvnrE39dmAHLmiuatmpS2lmY1kbdmIaZkMnpgk+nQG5Yb5vRaYZI+5Qf1l5P4OMYgm6UAa4
8rcqf2tGf6sY25nT3xq2hJW/teJ1b2F0LIPrm6uT9+T95+vLDHIssuBGr8PpJYhhWXX126iWuy2y
4etOAADDmE1FgOZj1O5oJKfCVoOjBjfGnLuNRfncbWauiAis89TtanvZIWTIuIfF1oA05iTqk7zz
j+YA1qk57Qst6x1MLc/fYKMFqFb2sr2hOpQTc0auY5aQjgDbGUS6tmdRLjKks+osMaFbgHQ7tS3A
gIpnkEXF6gFDTECsfzQH4T8tAQHIggporfSaCyLfzVpkXL61wa8DOW1z4GTYvx6Alwle/dbEAbKe
V3/OW/rx1fnV+xExEIuLgVGJwRaKwcnNiAwYc8sAQfCfAe5UIrBtInAxVgT4oiLArUoEtkIEWu/H
0QA8T8fTQfynVc9vU8+rP5NoAKaLi4GoxGALxWDUBmBjbhnIaQCvRGDbRGAsDcBsYRFAlQhshQh8
Ls7nY/lAzZIHkz7F0u2feGDBjzUhElHKOw/dmivOG6NqR8T9HSKilPCpgdNi6ogKm6bBU4wazGyQ
0XEigzP8c45yREXZeDVQtPkx1z7Q8AzkzCVB7tFtD+IMxRXO7DPOsCVwhlc4s3M4k8+Dw6vGGTKN
TLGcyNEVnClWJlSmKoBhTs8/H9Ri68hfVDi7t+PgFC9E50SDinTCfjUWvpl4yXJexvG0ecMD8FCC
R3M2V5PlyMzMpXLaVcdy3vUCYnCessnq1sXY9QKY9ScqgzzifSGHzKCzYxYm/eULyGhwa4GpymPO
C6imKk/AH46GF0aIBflaKan+tEFj0hyc0tM5UTRWmrFVsaT9nRhNGJlxYrQoPFFoJK7mC15XM6N3
wRHNQYEVi3QXC3jRWimpHF+ULEyHtv4xymLFWVfgVtr+ibD0lI41HnJUrLNY4JAjMtORZSY3q0OO
Ngo4zBw42KIHiJaSKs1Lno4bZo4bTKw25wo29nkFhVGtoNgRjHriaHjTcZbvtKM6habBmU3ZhISj
TMuRWPm242y5zYhtIRlDhm+scDPivAM2bTtiyixOkPm6G4/zpTYe38vD2EraTvQJm7iYVkn9mdr2
nuyE/iu5G8o2gV2pZuMVmFuWu4wFDInnesSZFEeTNwuy3Bc3XbsKtFQqXxRt2mCWWpJ9vRHbcBUm
VvkIc3gBY/Y7m2JekdVAdDNdANw8UQILpGm8upF8c1mCeB1jUseM1j7w2sGnAHRbBoCpKpWh0KpB
F45w4nKEE0B8TGi1inDOFuFcNApAUbHR3kz7MGHRDwMgs4GtkciiISw8pNtie4MAVcixr6WZ/24Y
K1X4iUO7+dO4cAcQWTZnXOQ8JdhZPJ8HOxFecd5V2OLHAh1GbAmgI0YFdHsCdLRgNmJZlaeooFO8
zuk0tKEFxJqrzZlVWPNjSVVx3MVCWIMrrNkTrOkPx9CFB0VweVBE0YuXh2Pw8HAMWW3OFa9ZZDhm
8VFcRBYfxcVsllFcbCGrGsbdCNxo/XPs7g7zbO+ArYF1/bja2GNLVvRmfZ//HbOse5513ZkY5Ou6
q72+tkMKfovBlkLf/nbW1JvJNHY1NqZfR6HXc9Pp3OnI6cgiodKbtRrAIpgLQLQHCUb+7LqW3d2x
0x0A0mmG6lztG08Vqte1X25vr1v9hr4Ow86b+zDu/u3we2w/tuvf08bofg3rgeweHoHxj7p/Y8DR
lJX9ltmcuBiY1b62oaV6UaehvYP33z3hd2Cg7tXBEa4qSgs4RNI1x3YvzjhWmns2xJuY0Td/hRQ6
bYdCVfOi1YeKOjTQK4nPsUWZh3460g5Lj0GjjqtKMhYKLdPr1rU3WS75+Ojf1Bjr28O3b9cyqkT6
Z7bOxFNoiadAv80UZn7d0dFd4hljUGYUYYx9RJglTuDgg0tqeN0S4/f+J/ne/0aDUaDpi4cqph5F
KGbQyXxoF/VnLVDS4GhQJZMZVBaqVPJHqOR3SE5rAWiD9oB6RdCqJze5Xg4tgKvsPm0gtlt23+XI
l5azHXbfWNTuQ7/xyu6/FshAG8d/2CVcERWuDOKK+tktXPEt5iC0Jf6EuSiuwI9R4cqr4cqVduX7
qp+VS6E1o55ja62TX5uX+azPyrGY07EY50TQXCxogxkNgpY7yG/5NSJpkvGR9h2ejo/S8ZS2TBaL
uD3oOah1r6O+XzZv048qzWzYZfyCWGTMMqUjXf9ByzNTswYpIUDgfXniCIqSzEg2N+W8wJ0FgVQ5
tBNVJOXCyCjO1d+s1H+16m+mPGUH1Z8vpv7mCAGo1P+Hqv8v4YPUw4QF6G/en1+8zXepR5Xurz+m
aM0YU9yFJaIEEzEzSBCRb9OTthKtQOKVQSJ3Fc5tJ+cHauF6FYcoxyGy2cm7FN9Eklr2lsQhxKJx
COg3s4pDvBa83ITQvb/aUdQOi0NySDVpYghckGgwslvgQi2fYWK41jaAC1sQXPJ+q8DlVcClNTXI
KSpPZ7VRDrX0ejejHMZiUQ5QlSrI+aogoCZoyCjIoSAjGXj//JcfHOHgSO1Ptz8RDmuhCIdqpSoM
+qMB4v8BvXdS9dpJAQA=
"""


def embedded_report():
    return gzip.decompress(base64.b64decode("".join(_EMBEDDED.split()))).decode("utf-8")


# ---------------------------------------------------------------- classification

# Ordered: first pattern that matches wins, so put specific before general.
CATEGORIES = [
    ("AZURE-SECRET",    r"AADSTS7000215"),
    ("BAD-CREDS",       r"INVALID_CREDENTIALS|Invalid credentials"),
    ("SSH-KEY",         r"SSH Key changed"),
    ("ZSCALER-API",     r"HTTPSConnectionPool|ConnectTimeoutError"),
    ("SSH-TIMEOUT",     r"Timeout exceeded"),
    ("NEVER-RETRIEVED", r"never received a retrieval status"),
    ("OK-STALE",        r"^Success"),
    ("LOGSERVER",       r"^Not applicable$"),
]

# Category -> (plain-English meaning, who has to act)
TRIAGE = {
    "AZURE-SECRET": (
        "Expired client secret on the Azure app registration",
        "Cloud Eng - rotate secret, then update it on the Azure management station",
    ),
    "BAD-CREDS": (
        "SSH connects but the service account is rejected",
        "Network Eng - supply current credentials for the FireMon service account",
    ),
    "SSH-KEY": (
        "Device SSH host key changed; FireMon will not accept it",
        "Enable 'Automatically Update SSH Keys' on the device, then force a retrieval",
    ),
    "SSH-TIMEOUT": (
        "SSH never answered - no banner, no auth prompt",
        "Network Eng - confirm mgmt IP is current and collector IP is permitted on vty",
    ),
    "ZSCALER-API": (
        "TCP connection to the Zscaler API timed out before authentication",
        "Collector owner - allow egress from the data collector to the API host on 443",
    ),
    "NEVER-RETRIEVED": (
        "Collector has never had a retrieval status for this device",
        "Onboarding was never completed - re-add the device",
    ),
    "OK-STALE": (
        "Last retrieval succeeded; nothing since",
        "No action on credentials - check schedule and change detection",
    ),
    "LOGSERVER": (
        "Log server - retrieval not applicable",
        "None",
    ),
    "NO-ERROR-TEXT": (
        "Failing but no retrieval error was recorded",
        "Force a manual retrieval to capture a current error",
    ),
    "NOT-LICENSED": (
        "Inactive or unlicensed - no health data collected",
        "Removal candidate if the device is decommissioned",
    ),
    "OTHER": ("Unrecognised retrieval error", "Review manually"),
}

# ---------------------------------------------------------------- parsing

DEVICE_RE = re.compile(r"^(?P<name>.+?)\s*\(ID:\s*(?P<id>\d+)\)\t(?P<rest>.*)$")
HEALTH_VALUES = {"Critical", "Healthy", "Inactive", "Unlicensed", "Warning"}

SECTION_HEADERS = {
    "Health Check Results", "GENERAL", "RETRIEVAL", "CHANGE DETECTION",
    "USAGE", "LAST RETRIEVAL", "LAST REVISION", "CHANGE MONITORING",
    "CHANGE DATA", "LOG MONITORING", "USAGE DATA", "DEVICE LICENSED",
    "DC GROUP ASSIGNED", "DEVICE UNLICENSED",
}


def _collect(lines, start_label, stop_labels):
    """Return the lines under start_label up to the next section header."""
    out = []
    grabbing = False
    for ln in lines:
        s = ln.strip()
        if s == start_label:
            grabbing = True
            continue
        if grabbing:
            if s in stop_labels or s in SECTION_HEADERS:
                break
            out.append(s)
    return " ".join(x for x in out if x).strip()


def parse_report(text):
    """text is the report contents, not a path."""
    raw = text.splitlines()

    devices = []
    i = 0
    while i < len(raw):
        m = DEVICE_RE.match(raw[i])
        if not m:
            i += 1
            continue

        cols = m.group("rest").split("\t")
        dev = {
            "id": int(m.group("id")),
            "name": m.group("name").strip(),
            "description": cols[0].strip() if len(cols) > 0 else "",
            "cluster": cols[1].strip() if len(cols) > 1 else "",
            "mgmt_ip": cols[2].strip() if len(cols) > 2 else "",
            "vendor": cols[3].strip() if len(cols) > 3 else "",
            "health": cols[4].strip() if len(cols) > 4 else "",
        }

        # Health usually lands on the following line; the trailing tab leaves it blank above.
        j = i + 1
        if not dev["health"] and j < len(raw) and raw[j].strip() in HEALTH_VALUES:
            dev["health"] = raw[j].strip()
            j += 1

        # Everything up to the next device row is this device's health block.
        block = []
        while j < len(raw) and not DEVICE_RE.match(raw[j]):
            block.append(raw[j])
            j += 1

        dev.update(parse_block(block))
        devices.append(dev)
        i = j

    return devices


def parse_block(lines):
    text = "\n".join(lines)
    out = {
        "licensed": "DEVICE LICENSED" in text,
        "collector": "",
        "last_retrieval": "",
        "retrieval_date": "",
        "rev_id": "", "rev_type": "", "rev_date": "",
        "rev_user": "", "rev_changes": "", "rev_result": "",
        "change_monitoring": "", "log_monitoring": "",
        "usage_data": "", "logged_conns": "",
        "ssh_target": "",
    }

    m = re.search(r"assigned to this device \(([^)]+)\)", text)
    if m:
        out["collector"] = m.group(1)

    out["last_retrieval"] = _collect(lines, "LAST RETRIEVAL", {"LAST REVISION"})

    m = re.search(r"[Ll]ast updated on (\d+/\d+/\d+) at ([\d:]+\s*[AP]M)",
                  out["last_retrieval"])
    if m:
        out["retrieval_date"] = f"{m.group(1)} {m.group(2)}"

    # pexpect dumps the exact ssh invocation - pull user@host out of it.
    m = re.search(r"'(-p)',\s*'(\d+)',\s*'([^']+@[\d.]+)'", out["last_retrieval"])
    if m:
        out["ssh_target"] = f"{m.group(3)}:{m.group(2)}"

    for key, label in [("rev_id", "Revision ID"), ("rev_type", "Type"),
                       ("rev_user", "User"), ("rev_changes", "Change Count"),
                       ("rev_result", "Result")]:
        m = re.search(rf"^{label}:\s*(.+)$", text, re.M)
        if m:
            out[key] = m.group(1).strip()

    m = re.search(r"^Date/Time:\s*(.+)$", text, re.M)
    if m:
        out["rev_date"] = m.group(1).strip()

    out["change_monitoring"] = _collect(lines, "CHANGE MONITORING", {"CHANGE DATA"})
    out["log_monitoring"] = _collect(lines, "LOG MONITORING", {"USAGE DATA"})
    out["usage_data"] = _collect(lines, "USAGE DATA", {"Logged Connections"})
    out["usage_data"] = re.sub(r"Logged Connections.*$", "", out["usage_data"]).strip()

    m = re.search(r"Logged Connections \(Last 24 hours\):\s*(\d+)", text)
    if m:
        out["logged_conns"] = int(m.group(1))

    return out


def classify(dev):
    if dev["health"] in ("Inactive", "Unlicensed"):
        return "NOT-LICENSED"
    lr = dev["last_retrieval"]
    if not lr:
        return "NO-ERROR-TEXT"
    for name, pattern in CATEGORIES:
        if re.search(pattern, lr):
            return name
    return "OTHER"


def days_since(datestr):
    """'6/29/26 8:44:34 PM' -> integer days, or '' when unparseable."""
    if not datestr:
        return ""
    m = re.match(r"(\d+)/(\d+)/(\d+)", datestr)
    if not m:
        return ""
    mo, day, yr = (int(x) for x in m.groups())
    yr += 2000 if yr < 100 else 0
    try:
        return (datetime.now() - datetime(yr, mo, day)).days
    except ValueError:
        return ""


# ---------------------------------------------------------------- workbook

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT, bold=True, size=14)
BODY = Font(name=FONT, size=10)
BOLD = Font(name=FONT, size=10, bold=True)
NOTE = Font(name=FONT, size=9, italic=True, color="595959")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SEVERITY_FILL = {
    "AZURE-SECRET": "FFC7CE",
    "BAD-CREDS": "FFC7CE",
    "SSH-KEY": "FFEB9C",
    "SSH-TIMEOUT": "FFC7CE",
    "ZSCALER-API": "FFC7CE",
    "NEVER-RETRIEVED": "FFEB9C",
    "NO-ERROR-TEXT": "FFEB9C",
    "OK-STALE": "C6EFCE",
    "LOGSERVER": "EDEDED",
    "NOT-LICENSED": "EDEDED",
}


def write_header(ws, headers, row=1):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_table(ws, name, ncols, nrows, header_row=1):
    if nrows == 0:
        return
    ref = f"A{header_row}:{get_column_letter(ncols)}{header_row + nrows}"
    t = Table(displayName=name, ref=ref)
    t.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(t)


def sheet_devices(wb, devices):
    ws = wb.create_sheet("All Devices")
    headers = [
        "Device ID", "Device Name", "Vendor", "Health", "Category",
        "Mgmt IP", "Cluster", "Data Collector", "Last Retrieval",
        "Days Stale", "Retrieval Detail", "SSH Target",
        "Revision Date", "Revision Result", "Revision User",
        "Change Monitoring", "Log Monitoring", "Usage Data", "Description",
    ]
    write_header(ws, headers)

    for r, d in enumerate(sorted(devices, key=lambda x: (x["vendor"], x["name"].lower())), 2):
        vals = [
            d["id"], d["name"], d["vendor"], d["health"], d["category"],
            d["mgmt_ip"], d["cluster"], d["collector"], d["retrieval_date"],
            days_since(d["retrieval_date"]), d["last_retrieval"][:500], d["ssh_target"],
            d["rev_date"], d["rev_result"], d["rev_user"],
            d["change_monitoring"], d["log_monitoring"][:300], d["usage_data"][:200],
            d["description"],
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = BODY
            cell.alignment = Alignment(vertical="top")
        fill = SEVERITY_FILL.get(d["category"])
        if fill:
            ws.cell(row=r, column=5).fill = PatternFill("solid", fgColor=fill)

    add_table(ws, "AllDevices", len(headers), len(devices))
    autosize(ws, [9, 42, 12, 10, 16, 15, 18, 30, 18, 10, 60, 24,
                  26, 20, 18, 18, 34, 30, 40])
    ws.auto_filter.ref = ws.dimensions
    return ws


def sheet_summary(wb, devices, src, report_date):
    ws = wb.create_sheet("Summary", 0)
    ws["A1"] = "FireMon Device Health - Summary"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Source: {src}"
    ws["A2"].font = NOTE
    ws["A3"] = f"Report generated by FireMon: {report_date or 'unknown'}"
    ws["A3"].font = NOTE
    ws["A4"] = f"Workbook built: {datetime.now():%Y-%m-%d %H:%M}"
    ws["A4"].font = NOTE
    ws["A5"] = ("Counts are COUNTIF formulas over 'All Devices' - edit the category "
                "there and every total below follows.")
    ws["A5"].font = NOTE

    n = len(devices)
    last = n + 1  # header occupies row 1 on All Devices

    row = 7
    ws.cell(row=row, column=1, value="BY HEALTH").font = BOLD
    row += 1
    write_header(ws, ["Health", "Devices"], row)
    ws.freeze_panes = None
    for h in ["Critical", "Healthy", "Inactive", "Unlicensed"]:
        row += 1
        ws.cell(row=row, column=1, value=h).font = BODY
        c = ws.cell(row=row, column=2,
                    value=f"=COUNTIF('All Devices'!$D$2:$D${last},A{row})")
        c.font = BODY
    row += 1
    ws.cell(row=row, column=1, value="Total").font = BOLD
    ws.cell(row=row, column=2, value=f"=COUNTA('All Devices'!$A$2:$A${last})").font = BOLD

    row += 2
    ws.cell(row=row, column=1, value="BY FAILURE CATEGORY").font = BOLD
    row += 1
    write_header(ws, ["Category", "Devices", "What it means", "Who acts"], row)
    cat_start = row + 1
    for cat in [c for c, _ in CATEGORIES] + ["NO-ERROR-TEXT", "NOT-LICENSED", "OTHER"]:
        row += 1
        ws.cell(row=row, column=1, value=cat).font = BODY
        ws.cell(row=row, column=2,
                value=f"=COUNTIF('All Devices'!$E$2:$E${last},A{row})").font = BODY
        meaning, owner = TRIAGE.get(cat, ("", ""))
        ws.cell(row=row, column=3, value=meaning).font = BODY
        ws.cell(row=row, column=4, value=owner).font = BODY
        fill = SEVERITY_FILL.get(cat)
        if fill:
            ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=fill)
    row += 1
    ws.cell(row=row, column=1, value="Total").font = BOLD
    ws.cell(row=row, column=2,
            value=f"=SUM(B{cat_start}:B{row - 1})").font = BOLD

    row += 2
    ws.cell(row=row, column=1, value="BY VENDOR").font = BOLD
    row += 1
    write_header(ws, ["Vendor", "Devices", "Licensed", "Critical"], row)
    for v in sorted({d["vendor"] for d in devices if d["vendor"]}):
        row += 1
        ws.cell(row=row, column=1, value=v).font = BODY
        ws.cell(row=row, column=2,
                value=f"=COUNTIF('All Devices'!$C$2:$C${last},A{row})").font = BODY
        ws.cell(row=row, column=3,
                value=(f"=COUNTIFS('All Devices'!$C$2:$C${last},A{row},"
                       f"'All Devices'!$D$2:$D${last},\"Critical\")"
                       f"+COUNTIFS('All Devices'!$C$2:$C${last},A{row},"
                       f"'All Devices'!$D$2:$D${last},\"Healthy\")")).font = BODY
        ws.cell(row=row, column=4,
                value=(f"=COUNTIFS('All Devices'!$C$2:$C${last},A{row},"
                       f"'All Devices'!$D$2:$D${last},\"Critical\")")).font = BODY

    autosize(ws, [30, 12, 58, 62])
    return ws


def sheet_actions(wb, devices):
    ws = wb.create_sheet("Action List")
    ws["A1"] = "Devices needing action, grouped by root cause"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = ("Excludes log servers, inactive/unlicensed devices, and devices whose "
                "last retrieval succeeded.")
    ws["A2"].font = NOTE

    actionable = [c for c, _ in CATEGORIES if c not in ("OK-STALE", "LOGSERVER")]
    actionable += ["NO-ERROR-TEXT", "OTHER"]

    row = 4
    for cat in actionable:
        rows = [d for d in devices if d["category"] == cat]
        if not rows:
            continue
        meaning, owner = TRIAGE.get(cat, ("", ""))
        ws.cell(row=row, column=1, value=f"{cat}  ({len(rows)})").font = Font(
            name=FONT, bold=True, size=11)
        fill = SEVERITY_FILL.get(cat)
        if fill:
            ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=fill)
        row += 1
        ws.cell(row=row, column=1, value=meaning).font = NOTE
        row += 1
        ws.cell(row=row, column=1, value=f"Owner: {owner}").font = NOTE
        row += 1

        write_header(ws, ["Device ID", "Device Name", "Vendor", "Mgmt IP",
                          "SSH Target", "Last Retrieval", "Days Stale",
                          "Data Collector"], row)
        ws.freeze_panes = None
        row += 1
        for d in sorted(rows, key=lambda x: x["name"].lower()):
            vals = [d["id"], d["name"], d["vendor"], d["mgmt_ip"], d["ssh_target"],
                    d["retrieval_date"], days_since(d["retrieval_date"]), d["collector"]]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.font = BODY
                cell.border = BOX
            row += 1
        row += 2

    autosize(ws, [10, 46, 12, 17, 26, 20, 11, 34])
    return ws


def sheet_licensing(wb, devices):
    ws = wb.create_sheet("Licensing")
    ws["A1"] = "Licensing position"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = ("Derived from the Device Health Report. Confirm against a clean run of the "
                "Device Inventory Report before quoting externally.")
    ws["A2"].font = NOTE

    lic = [d for d in devices if d["health"] in ("Critical", "Healthy")]
    unl = [d for d in devices if d["health"] == "Unlicensed"]
    ina = [d for d in devices if d["health"] == "Inactive"]

    row = 4
    write_header(ws, ["Vendor", "Licensed", "Unlicensed", "Inactive", "Total"], row)
    ws.freeze_panes = None
    start = row + 1
    for v in sorted({d["vendor"] for d in devices if d["vendor"]}):
        row += 1
        ws.cell(row=row, column=1, value=v).font = BODY
        for col, subset in ((2, lic), (3, unl), (4, ina)):
            ws.cell(row=row, column=col,
                    value=sum(1 for d in subset if d["vendor"] == v)).font = BODY
        ws.cell(row=row, column=5, value=f"=SUM(B{row}:D{row})").font = BODY
    row += 1
    ws.cell(row=row, column=1, value="Total").font = BOLD
    for col in range(2, 6):
        L = get_column_letter(col)
        ws.cell(row=row, column=col, value=f"=SUM({L}{start}:{L}{row - 1})").font = BOLD

    row += 2
    ws.cell(row=row, column=1,
            value="Devices not consuming a license (removal candidates)").font = BOLD
    row += 1
    write_header(ws, ["Device ID", "Device Name", "Vendor", "Health",
                      "Mgmt IP", "Cluster", "Description"], row)
    ws.freeze_panes = None
    row += 1
    for d in sorted(unl + ina, key=lambda x: (x["vendor"], x["name"].lower())):
        vals = [d["id"], d["name"], d["vendor"], d["health"],
                d["mgmt_ip"], d["cluster"], d["description"]]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font = BODY
            cell.border = BOX
        row += 1

    autosize(ws, [10, 46, 12, 12, 17, 22, 48])
    return ws


def build(devices, src, report_date, out_path):
    wb = Workbook()
    wb.remove(wb.active)
    sheet_devices(wb, devices)
    sheet_actions(wb, devices)
    sheet_licensing(wb, devices)
    sheet_summary(wb, devices, src, report_date)
    wb.move_sheet("Summary", offset=-3)
    wb.save(out_path)


def main():
    if len(sys.argv) > 1:
        src = sys.argv[1]
        with open(src, "r", encoding="utf-8", errors="replace") as fh:
            text = fh.read()
        out = sys.argv[2] if len(sys.argv) > 2 else "firemon_device_health.xlsx"
    else:
        src = "embedded report (21 Aug 2026)"
        text = embedded_report()
        out = "firemon_device_health.xlsx"
        print("No input file given - using the embedded 21 Aug 2026 report.")
        print("Pass a newer export as an argument to use that instead.\n")

    devices = parse_report(text)
    if not devices:
        print("No devices parsed. Is this the Device Health Report text export?")
        sys.exit(1)

    for d in devices:
        d["category"] = classify(d)

    report_date = ""
    for ln in text.splitlines()[:5]:
        if re.match(r"^[A-Z][a-z]+ \d+, \d{4}", ln.strip()):
            report_date = ln.strip()
            break

    build(devices, src, report_date, out)

    print(f"Parsed {len(devices)} devices -> {out}\n")
    for cat, n in Counter(d["category"] for d in devices).most_common():
        print(f"  {n:4d}  {cat:16s} {TRIAGE.get(cat, ('', ''))[0]}")
    print()
    for h, n in Counter(d["health"] for d in devices).most_common():
        print(f"  {n:4d}  {h}")


if __name__ == "__main__":
    main()
