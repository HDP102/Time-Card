"""
notify_rule_owners.py
=====================
Workstream 1 — Rule Recertification Owner Notifications
PG&E | NPS Automation

Reads the firewall rule spreadsheet, resolves each rule's owner from the AMPS data,
groups rules by owner, and contacts them with three options:
  A) Recertify   B) Clean up / Remove   C) Review with the team

Every owner receives an Excel attachment with a Decision dropdown — the response
format is uniform so replies can be tracked and audited consistently.

RUN MODES (set RUN_MODE in CONFIG):
  notify    Email each owner their rules, plus an optional companion Teams card.
  remind    Chase owners who have not responded, with a days-outstanding count.
  announce  One heads-up card to the team channel before a wave goes out.
  digest    One status card to the team channel. Contacts no owners.

Filters (NOTIFY_ONLY / EXCLUDE_OWNERS / NOTIFY_DEVICES) apply to every mode, and
in dry run, test, and real sends alike.

Teams notes:
  - Webhook URLs must come from the Workflows / Power Automate app. Classic Office
    365 connector webhooks were retired by Microsoft in May 2026.
  - Owner-directed cards need TEAMS_USER_WEBHOOK, a flow that reads a "recipient"
    field and DMs that person. Without it, owner cards fall back to the channel.
  - Cards never carry rule data — they point the owner at the email.

Usage:
  py -m pip install openpyxl requests
  py notify_rule_owners.py
"""

import ast
import csv
import io
import json
import logging
import smtplib
import re
import time
from datetime import datetime, timezone
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from pathlib import Path
from urllib.parse import quote

import requests
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────

# ── What this run does ───────────────────────────────────────────────
# RUN_MODE picks the job. Filters apply to every mode.
#   "notify"   — the main run. Emails each owner their rules as an Excel attachment,
#                and optionally posts a companion Teams card telling them to check email.
#   "remind"   — chases owners who have not responded yet. Same audience minus anyone
#                listed in RESPONDED_FILE, with reminder wording and a days-outstanding
#                count. Requires a prior notify run so there is history to read.
#   "announce" — one heads-up card to the team channel before a notify wave goes out.
#                Contacts no owners. Cuts down "is this email real?" replies.
#   "digest"   — one status card to the team channel: notified / responded / outstanding,
#                oldest outstanding, biggest holdouts. Contacts no owners.
RUN_MODE             = "notify"

EXCEL_PATH           = "firewall_rule_report.xlsx"  # Path to the firewall rule spreadsheet
NOTIFY_EMAIL         = True                          # Send email notifications
NOTIFY_TEAMS         = False                         # Send Teams notifications
DRY_RUN              = True                          # True = print only, don't send

SENDER_EMAIL         = "corpid@pge.com"              # Your CorpID@pge.com
SMTP_HOST            = "mailhost"                    # PG&E internal mailhost
SMTP_PORT            = 25

# ── Teams ────────────────────────────────────────────────────────────
# Webhook URLs must come from the Workflows / Power Automate app. Classic Office 365
# connector webhooks were retired by Microsoft in May 2026 and no longer deliver.
#
# TEAMS_WEBHOOK — the team channel. Used for "announce" and "digest" cards, and as the
#   fallback for owner cards if no per-user flow is configured.
TEAMS_WEBHOOK        = "<YOUR_TEAMS_CHANNEL_WEBHOOK>"

# TEAMS_USER_WEBHOOK — a Power Automate flow that DMs an individual. The flow must read
#   a "recipient" field from the request body and post to that person's chat. Owner cards
#   go here when set. Leave "" to fall back to the channel webhook above.
#   An owner's UPN is their CorpID address, which the script already has.
TEAMS_USER_WEBHOOK   = ""

# TEAMS_WEBHOOK_TEST — used instead of the above whenever TEST_MODE = True. Point it at
#   your own chat or a private channel. Leave "" to reuse the real destination.
TEAMS_WEBHOOK_TEST   = ""

# Seconds between Teams posts. Webhooks throttle a tight loop — keep at 1 or higher.
TEAMS_DELAY_SECONDS  = 1

LOG_FILE             = f"notifications_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

# ── Response tracking (used by "remind" and "digest") ────────────────
# RESPONDED_FILE — CSV of owners who have already replied, so reminders skip them.
#   One CorpID per line, or a "corpid" column. Maintained by hand for now; swap in a
#   real source once response tracking is settled.
RESPONDED_FILE       = "responded.csv"

# RESPONSES_DIR — drop replied-to spreadsheets here. Both this folder and the file
#   above are created automatically after the first real send.
RESPONSES_DIR        = "responses"

# LOGS_DIR — run logs live here. Reminders and digests read them for history.
#   To throw away a bad run, delete its log or move it into LOGS_DIR/discarded —
#   anything in that subfolder is ignored, so the run stops counting as notified.
LOGS_DIR             = "logs"

# REMIND_AFTER_DAYS — only chase owners notified at least this many days ago.
#   Notification dates are read from the notifications_*.csv logs of previous runs.
REMIND_AFTER_DAYS    = 14

# ── Mode ─────────────────────────────────────────────────────────────
# TEST MODE — same filters, but email goes to SENDER_EMAIL and Teams cards go to
# TEAMS_WEBHOOK_TEST. Cards are labelled as tests. Scope it with the filters below.
TEST_MODE            = False

# TEST_LIMIT — hard cap on how many emails a test run sends you, so an unfiltered
#   test cannot flood your inbox. 1 = one sample owner, 0 = no cap (send them all).
#   Combine with NOTIFY_ONLY to choose exactly whose email you want to see.
TEST_LIMIT           = 1

# ── FILTERS — scope who gets notified. Apply in ALL modes: dry run, test, real send.
#    All three stack: an owner must pass every active filter to be notified.
#    Leave a list empty to turn that filter off.

# NOTIFY_ONLY — whitelist. ONLY these CorpIDs get notified. [] = notify everyone.
#   Example: NOTIFY_ONLY = ["ABC1", "XYZ2"]
NOTIFY_ONLY          = []

# EXCLUDE_OWNERS — blacklist. These CorpIDs are ALWAYS skipped, even if whitelisted.
#   Example: EXCLUDE_OWNERS = ["ABC1"]
EXCLUDE_OWNERS       = []

# NOTIFY_DEVICES — only owners who have rules on these devices/sites.
#   Partial, case-insensitive match against the tab name, Device Name, and Policy Name.
#   A short site name matches all its variants, e.g. "Site_A" matches
#   "Site_A Sub", "Site_A_Station", etc.
#   An owner's rule list is also trimmed to just the matching devices, so they only
#   see rules for the site you are working on.
#   Example: NOTIFY_DEVICES = ["Site_Name_A", "Site_Name_B"]
NOTIFY_DEVICES       = []

# Tags to SKIP — base rules and standard exceptions don't need owner notifications
SKIP_TAGS            = {"BaseRule", "ToolsRule", "ToolsRules"}

# Shadowing: skip fully shadowed rules (already handled by decommission workstream)
SKIP_SHADOWED        = True
# ─────────────────────────────────────────────

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

DECISION_OPTIONS = "A) Recertify  |  B) Clean up / Remove  |  C) Review with team"


# ── Owner parsing ──────────────────────────────────────────────────────────────

def parse_owner_string(raw):
    owners = []
    if not raw or "Skip AMPS" in raw or "No APPID" in raw:
        return owners
    for line in raw.strip().split("\n"):
        line = line.strip()
        if not line or "=" not in line:
            continue
        try:
            ip_part, dict_part = line.split("=", 1)
            ip = ip_part.strip()
            data = ast.literal_eval(dict_part.strip())
            for app_id, hostname_map in data.items():
                for hostname, inner in hostname_map.items():
                    for app_key, fields in inner.items():
                        if not isinstance(fields, dict):
                            continue
                        for role, value in fields.items():
                            if not value or not isinstance(value, str):
                                continue
                            match = re.match(r"(.+?)\s*\(([A-Z0-9]+)\)", value)
                            if match:
                                name   = match.group(1).strip()
                                corpid = match.group(2).strip()
                                email  = f"{corpid}@pge.com"
                                owners.append({
                                    "email":    email,
                                    "name":     name,
                                    "corpid":   corpid,
                                    "role":     role,
                                    "app_id":   app_id,
                                    "ip":       ip,
                                    "hostname": hostname,
                                })
        except Exception:
            continue
    return owners


def extract_primary_owner(owners):
    priority = ["Client Owner", "Cyber Owner", "IT SME", "IT Lead", "IT SME backup"]
    for role in priority:
        for o in owners:
            if o["role"] == role:
                return o
    return owners[0] if owners else None


# ── Excel reading ──────────────────────────────────────────────────────────────

def load_rules(excel_path):
    log.info(f"Loading spreadsheet: {excel_path}")
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    rules = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = None

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = list(row)
                col = {h: idx for idx, h in enumerate(headers) if h}
                continue
            if not any(row):
                continue

            def g(field):
                idx = col.get(field)
                return row[idx] if idx is not None else None

            tags = str(g("Tags") or "")
            if any(t in tags for t in SKIP_TAGS):
                continue

            shadow = str(g("Shadowing Status") or "")
            if SKIP_SHADOWED and shadow == "FULLY_SHADOWED":
                continue

            rules.append({
                "sheet":        sheet_name,
                "device_name":  g("Device Name"),
                "policy_name":  g("Policy Name"),
                "rule_name":    g("ID on Device"),
                "source":       g("Source"),
                "destination":  g("Destination"),
                "service":      g("Service"),
                "application":  g("Application"),
                "action":       g("Action"),
                "tags":         tags,
                "last_hit":     g("Last Hit"),
                "shadowing":    shadow,
                "ticket_id":    g("Ticket ID"),
                "app_name":     g("Application Name"),
                "disabled":     g("Disabled"),
                "src_owners_raw": g("Source AMPS owners list"),
                "dst_owners_raw": g("Destination AMPS owners list"),
                "business_justification": g("Policy Name + tag + FER close date + Business Justification"),
            })

    total_sheets = len(wb.sheetnames)
    wb.close()
    log.info(f"Loaded {len(rules)} rules across {total_sheets} devices")
    return rules


# ── Owner aggregation ──────────────────────────────────────────────────────────

def build_owner_rule_map(rules):
    owner_map = {}
    for rule in rules:
        all_owners = []
        all_owners += parse_owner_string(rule.get("src_owners_raw", "") or "")
        all_owners += parse_owner_string(rule.get("dst_owners_raw", "") or "")
        owner = extract_primary_owner(all_owners)
        if not owner:
            continue
        email = owner["email"]
        if email not in owner_map:
            owner_map[email] = {"owner": owner, "rules": []}
        owner_map[email]["rules"].append(rule)
    return owner_map


# ── Filters ────────────────────────────────────────────────────────────────────

def rule_matches_devices(rule, device_filters):
    """True if a rule belongs to any of the requested devices/sites (partial, case-insensitive)."""
    haystack = " ".join(str(rule.get(f) or "") for f in ("sheet", "device_name", "policy_name")).lower()
    return any(d.strip().lower() in haystack for d in device_filters if d.strip())


def apply_filters(owner_map):
    """
    Narrow the owner map down using NOTIFY_ONLY / EXCLUDE_OWNERS / NOTIFY_DEVICES.
    An owner must pass every active filter. Returns a new map — the original is untouched.
    """
    whitelist = {c.strip().upper() for c in NOTIFY_ONLY if c.strip()}
    blacklist = {c.strip().upper() for c in EXCLUDE_OWNERS if c.strip()}
    devices   = [d for d in NOTIFY_DEVICES if d.strip()]

    if not (whitelist or blacklist or devices):
        return owner_map

    filtered = {}
    for email, data in owner_map.items():
        owner  = data["owner"]
        rules  = data["rules"]
        corpid = (owner.get("corpid") or "").strip().upper()

        # 1. Whitelist — if set, only these CorpIDs pass
        if whitelist and corpid not in whitelist:
            continue

        # 2. Blacklist — always wins over the whitelist
        if corpid in blacklist:
            continue

        # 3. Device filter — owner must have at least one rule on a matching device,
        #    and their rule list is trimmed to only those rules
        if devices:
            rules = [r for r in rules if rule_matches_devices(r, devices)]
            if not rules:
                continue

        filtered[email] = {"owner": owner, "rules": rules}

    return filtered


def log_active_filters():
    """Print which filters are on so there is no doubt about who is in scope."""
    active = []
    if NOTIFY_ONLY:
        active.append(f"NOTIFY_ONLY: {NOTIFY_ONLY}")
    if EXCLUDE_OWNERS:
        active.append(f"EXCLUDE_OWNERS: {EXCLUDE_OWNERS}")
    if NOTIFY_DEVICES:
        active.append(f"NOTIFY_DEVICES: {NOTIFY_DEVICES}")

    if not active:
        log.info("  No filters active — all owners in scope")
        return
    for line in active:
        log.info(f"  {line}")


# ── Ticket ID cleaning ─────────────────────────────────────────────────────────

def clean_ticket_id(raw_ticket):
    ticket_lines = []
    for t in str(raw_ticket or "").strip().split("\n"):
        t = t.strip()
        if t and t not in ("0", "0.0"):
            try:
                ticket_lines.append(str(int(float(t))) if "." in t else t)
            except ValueError:
                ticket_lines.append(t)
    return ", ".join(ticket_lines) if ticket_lines else "N/A"


def truncate(text, length=60):
    text = str(text or "")
    return (text[:length] + "...") if len(text) > length else text


# ── Excel attachment builder ───────────────────────────────────────────────────

def build_excel_attachment(owner, rules):
    """Build an Excel file in memory with a Decision dropdown column for the owner to fill in."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Rules for Review"

    blue_fill  = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    light_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # Row 1 — instruction banner
    ws.merge_cells("A1:G1")
    ws["A1"] = "Firewall Rule Recertification — Please fill in the Decision column for each rule using the dropdown, then reply with this completed file."
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=11)
    ws["A1"].fill = blue_fill
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32

    # Row 2 — A/B/C legend
    ws.merge_cells("A2:G2")
    ws["A2"] = (
        "A) Recertify — Rule is still needed and should remain active     |     "
        "B) Clean up / Remove — Rule is no longer needed and can be disabled/deleted     |     "
        "C) Review with team — Unsure, would like to discuss with the NPS Automation team"
    )
    ws["A2"].font = Font(bold=True, color="003366", size=10)
    ws["A2"].fill = light_fill
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 28

    # Row 3 — column headers
    headers = ["Device", "Rule Name", "Source", "Destination", "Last Hit", "Ticket ID", "Decision (A / B / C)"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = blue_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="left")

    # Data rows starting at row 4
    for row_idx, r in enumerate(rules, 4):
        ws.cell(row=row_idx, column=1, value=str(r.get("device_name") or ""))
        ws.cell(row=row_idx, column=2, value=str(r.get("rule_name") or ""))
        ws.cell(row=row_idx, column=3, value=str(r.get("source") or ""))
        ws.cell(row=row_idx, column=4, value=str(r.get("destination") or ""))
        ws.cell(row=row_idx, column=5, value=str(r.get("last_hit") or "Never"))
        ws.cell(row=row_idx, column=6, value=clean_ticket_id(r.get("ticket_id")))
        cell = ws.cell(row=row_idx, column=7, value="")
        cell.alignment = Alignment(horizontal="center")

    # Dropdown on Decision column — rows 4 to end of data
    last_data_row = len(rules) + 3
    dv = DataValidation(
        type="list",
        formula1='"A) Recertify,B) Clean up / Remove,C) Review with team"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.sqref = f"G4:G{last_data_row}"
    ws.add_data_validation(dv)

    # Column widths
    col_widths = [25, 30, 40, 40, 15, 15, 28]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


# ── Email formatting ───────────────────────────────────────────────────────────

def get_first_name(owner):
    parts = owner["name"].split(",")
    return parts[1].strip().split()[0] if len(parts) > 1 else parts[0].strip()


def format_email_attachment(owner, rules):
    """Email body when rules are in an attachment (more than 8 rules)."""
    name = get_first_name(owner)

    return f"""
    <html><body style='font-family:Arial,sans-serif;color:#333'>
    <p>Hi {name},</p>
    <p>As part of PG&E's annual firewall rule recertification process, the following
    firewall rules have been identified as requiring your review. These rules are
    associated with applications or assets you own.</p>
    <p>Your full rule list (<strong>{len(rules)} rule{'s' if len(rules) != 1 else ''}</strong>)
    is attached as an Excel file.</p>
    <p><strong>Please open the attached file, fill in the Decision column for each rule,
    and reply to this email with the completed attachment:</strong></p>
    <ol>
      <li><strong>A) Recertify</strong> — The rule is still needed and should remain active.</li>
      <li><strong>B) Clean up / Remove</strong> — The rule is no longer needed and can be disabled/deleted.</li>
      <li><strong>C) Review with the team</strong> — You are unsure and would like to discuss with the NPS Automation team.</li>
    </ol>
    <p>If you have any questions or would like to schedule a review session,
    please contact the NPS Automation team directly.</p>
    <p>Thank you,<br><strong>NPS Automation Team</strong><br>PG&E Network &amp; Platform Security</p>
    <p style='font-size:11px;color:#888'>This is an automated notification from the NPS firewall recertification process.
    Rule count in this notification: {len(rules)}</p>
    </body></html>
    """


def format_email_reminder(owner, rules, days_outstanding=None):
    """Chase email. Assumes they have already read the original — no process explainer."""
    name   = get_first_name(owner)
    count  = len(rules)
    waited = (f" It has been <strong>{days_outstanding} days</strong> since that request."
              if days_outstanding else "")

    return f"""
    <html><body style='font-family:Arial,sans-serif;color:#333'>
    <p>Hi {name},</p>
    <p>We have not yet received your recertification decisions for
    <strong>{count} firewall rule{'s' if count != 1 else ''}</strong>
    associated with applications or assets you own.{waited}</p>
    <p>The same rule list is attached again so you do not need to look for the original
    email. <strong>Please fill in the Decision column for each rule and reply with the
    completed file:</strong></p>
    <ol>
      <li><strong>A) Recertify</strong> — The rule is still needed and should remain active.</li>
      <li><strong>B) Clean up / Remove</strong> — The rule is no longer needed and can be disabled/deleted.</li>
      <li><strong>C) Review with the team</strong> — You are unsure and would like to discuss with the NPS Automation team.</li>
    </ol>
    <p>If you have already replied, or if these rules belong to someone else now,
    let us know and we will update our records.</p>
    <p>If you would prefer to walk through the list together, contact the NPS Automation
    team and we will set up a review session.</p>
    <p>Thank you,<br><strong>NPS Automation Team</strong><br>PG&E Network &amp; Platform Security</p>
    <p style='font-size:11px;color:#888'>This is an automated reminder from the NPS firewall
    recertification process. Rule count in this notification: {count}</p>
    </body></html>
    """


# ── Sending ────────────────────────────────────────────────────────────────────

def send_email(to_email, owner, rules, dry_run=True, reminder=False, days_outstanding=None):
    """Every owner gets an Excel attachment — the response format is uniform by design."""
    rule_count = len(rules)
    plural     = "s" if rule_count != 1 else ""

    if reminder:
        subject   = f"Reminder: Firewall Rule Recertification Still Outstanding ({rule_count} rule{plural})"
        html_body = format_email_reminder(owner, rules, days_outstanding)
    else:
        subject   = f"Action Required: Firewall Rule Recertification ({rule_count} rule{plural})"
        html_body = format_email_attachment(owner, rules)

    if dry_run:
        log.info(f"[DRY RUN] Would email: {to_email} | {rule_count} rules | Subject: {subject}")
        return True

    msg = MIMEMultipart("mixed")
    msg["Subject"] = subject
    msg["From"]    = SENDER_EMAIL
    msg["To"]      = to_email
    msg.attach(MIMEText(html_body, "html"))

    excel_bytes = build_excel_attachment(owner, rules)
    part = MIMEBase("application", "octet-stream")
    part.set_payload(excel_bytes)
    encoders.encode_base64(part)
    filename = f"Firewall_Rules_Recertification_{owner['corpid']}.xlsx"
    part.add_header("Content-Disposition", f"attachment; filename={filename}")
    msg.attach(part)

    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.sendmail(SENDER_EMAIL, [to_email], msg.as_string())
        log.info(f"Email sent: {to_email} | {rule_count} rules")
        return True
    except Exception as e:
        log.error(f"Email failed to {to_email}: {e}")
        return False


# ── Response tracking ──────────────────────────────────────────────────────────

def load_responded():
    """CorpIDs that have already replied. One per line, or a 'corpid' column."""
    path = Path(RESPONDED_FILE)
    if not path.exists():
        log.info(f"No response file at {RESPONDED_FILE} — treating every owner as outstanding")
        return set()

    responded = set()
    with open(path, newline="", encoding="utf-8-sig") as f:
        for row in csv.reader(f):
            if not row:
                continue
            val = row[0].strip()
            if not val or val.lower() == "corpid":
                continue
            responded.add(val.upper())
    log.info(f"Loaded {len(responded)} responded owner(s) from {RESPONDED_FILE}")
    return responded


def load_notification_history():
    """First-notified timestamp per CorpID, read from previous notifications_*.csv logs."""
    history = {}
    # glob is non-recursive, so anything moved into LOGS_DIR/discarded is ignored
    candidates = list(Path(LOGS_DIR).glob("notifications_*.csv"))
    candidates += list(Path(".").glob("notifications_*.csv"))   # logs from older runs
    for path in sorted(set(candidates)):
        if path.name == LOG_FILE:
            continue
        try:
            with open(path, newline="", encoding="utf-8-sig") as f:
                for row in csv.DictReader(f):
                    corpid = (row.get("corpid") or "").strip().upper()
                    stamp  = (row.get("timestamp") or "").strip()
                    if not corpid or not stamp:
                        continue
                    # rehearsals are not notifications — ignore dry runs and test sends
                    if str(row.get("email_sent")).lower() != "true":
                        continue
                    if str(row.get("dry_run", "")).lower() == "true":
                        continue
                    if str(row.get("test_mode", "")).lower() == "true":
                        continue
                    try:
                        when = datetime.fromisoformat(stamp)
                    except ValueError:
                        continue
                    if corpid not in history or when < history[corpid]:
                        history[corpid] = when
        except Exception as e:
            log.warning(f"Could not read {path.name}: {e}")

    if history:
        log.info(f"Read notification history for {len(history)} owner(s) from previous run logs")
    else:
        log.info("No previous notification logs found — no history to age reminders against")
    return history


def days_since(when):
    if not when:
        return None
    if when.tzinfo is None:
        when = when.replace(tzinfo=timezone.utc)
    return (datetime.now(timezone.utc) - when).days


def build_stats(owner_map, responded, history):
    """Roll-up used by the digest card."""
    outstanding = [(d["owner"], d["rules"]) for d in owner_map.values()
                   if d["owner"]["corpid"].upper() not in responded]
    ages = [days_since(history.get(o["corpid"].upper())) for o, _ in outstanding]
    ages = [a for a in ages if a is not None]

    holdouts = sorted(outstanding, key=lambda x: len(x[1]), reverse=True)[:5]
    return {
        "notified":          len(owner_map),
        "responded":         len(owner_map) - len(outstanding),
        "outstanding":       len(outstanding),
        "rules_outstanding": sum(len(r) for _, r in outstanding),
        "oldest_days":       max(ages) if ages else None,
        "holdouts":          [(o["corpid"], len(r), days_since(history.get(o["corpid"].upper())))
                              for o, r in holdouts],
    }


# ── Teams cards ────────────────────────────────────────────────────────────────

DECISION_HELP = ("✅ **A) Recertify** — Rule is still needed\n\n"
                 "🗑️ **B) Clean up / Remove** — Rule is no longer needed\n\n"
                 "💬 **C) Review with the team** — Need to discuss")


def _card(body, actions=None):
    """Wrap card body blocks in the Teams adaptive-card envelope."""
    content = {
        "$schema": "http://adaptivecards.io/schemas/adaptive-card.json",
        "type": "AdaptiveCard",
        "version": "1.4",
        "body": body,
    }
    if actions:
        content["actions"] = actions
    return {"type": "message",
            "attachments": [{"contentType": "application/vnd.microsoft.card.adaptive",
                             "content": content}]}


def _test_banner(owner=None):
    who = f" Generated for {owner['name']} ({owner['corpid']})." if owner else ""
    return {"type": "TextBlock",
            "text": f"🧪 **TEST — not an official notification.**{who}",
            "wrap": True, "color": "Warning", "weight": "Bolder"}


def _reply_action(owner, rule_count, reminder=False):
    """One-tap mailto reply. Avoids a spreadsheet round trip for simple answers."""
    subject = quote(f"Firewall Rule Recertification - {owner['corpid']} ({rule_count} rules)")
    return [{"type": "Action.OpenUrl", "title": "Reply by email",
             "url": f"mailto:{SENDER_EMAIL}?subject={subject}"}]


def build_owner_card(owner, rules, test_mode=False, reminder=False, days_outstanding=None):
    """Companion card (notify) or chase card (remind). Never carries the rule data itself."""
    name = get_first_name(owner)
    body = []
    if test_mode:
        body.append(_test_banner(owner))

    if reminder:
        waited = f" It has been {days_outstanding} days since that notice." if days_outstanding else ""
        body += [
            {"type": "TextBlock", "text": "⏰ Reminder: Firewall Rule Recertification Outstanding",
             "weight": "Bolder", "size": "Medium", "color": "Warning"},
            {"type": "TextBlock",
             "text": f"Hi **{name}**, we have not yet received your decisions for "
                     f"**{len(rules)} firewall rule{'s' if len(rules) != 1 else ''}**.{waited}",
             "wrap": True},
            {"type": "TextBlock",
             "text": "The original email has an Excel attachment with a Decision column. "
                     "Please complete it and reply, or let us know if you need a review session.",
             "wrap": True, "spacing": "Small"},
        ]
    else:
        body += [
            {"type": "TextBlock", "text": "🔒 Firewall Rule Recertification Required",
             "weight": "Bolder", "size": "Medium", "color": "Accent"},
            {"type": "TextBlock",
             "text": f"Hi **{name}**, you have **{len(rules)} firewall rule"
                     f"{'s' if len(rules) != 1 else ''}** awaiting recertification.",
             "wrap": True},
            {"type": "TextBlock",
             "text": f"📧 **Check your email.** A message from {SENDER_EMAIL} has an Excel "
                     f"attachment listing your rules, with a dropdown to record a decision "
                     f"for each one. This card is confirmation that the email is legitimate.",
             "wrap": True, "spacing": "Small"},
        ]

    body += [
        {"type": "TextBlock", "text": "**Decision Options:**", "weight": "Bolder", "spacing": "Medium"},
        {"type": "TextBlock", "text": DECISION_HELP, "wrap": True, "spacing": "Small"},
        {"type": "TextBlock",
         "text": "Reply to the email with the completed attachment, or contact the NPS Automation team.",
         "wrap": True, "spacing": "Medium", "isSubtle": True},
    ]
    return _card(body, _reply_action(owner, len(rules), reminder))


def build_announcement_card(owner_count, rule_count, test_mode=False):
    """Heads-up to the team channel before a notify wave goes out."""
    body = []
    if test_mode:
        body.append(_test_banner())
    body += [
        {"type": "TextBlock", "text": "📢 Firewall Rule Recertification — Notifications Going Out",
         "weight": "Bolder", "size": "Medium", "color": "Accent"},
        {"type": "TextBlock",
         "text": f"**{owner_count} rule owner{'s' if owner_count != 1 else ''}** are about to receive "
                 f"a recertification request covering **{rule_count} rules**.",
         "wrap": True},
        {"type": "TextBlock",
         "text": f"**What owners will see:** an email from {SENDER_EMAIL} with an Excel attachment "
                 f"listing their rules and a Decision dropdown (A / B / C).",
         "wrap": True, "spacing": "Small"},
        {"type": "TextBlock",
         "text": "**This is a legitimate internal request.** If someone asks whether it is phishing, "
                 "it is not — please point them here.",
         "wrap": True, "spacing": "Small", "weight": "Bolder"},
        {"type": "TextBlock", "text": DECISION_HELP, "wrap": True, "spacing": "Medium"},
        {"type": "TextBlock",
         "text": "Questions or escalations go to the NPS Automation team.",
         "wrap": True, "spacing": "Medium", "isSubtle": True},
    ]
    return _card(body)


def build_digest_card(stats, test_mode=False):
    """Status roll-up to the team channel. Contacts no owners."""
    body = []
    if test_mode:
        body.append(_test_banner())

    facts = [
        {"title": "Owners notified", "value": str(stats["notified"])},
        {"title": "Responded", "value": str(stats["responded"])},
        {"title": "Outstanding", "value": str(stats["outstanding"])},
        {"title": "Rules outstanding", "value": str(stats["rules_outstanding"])},
    ]
    if stats.get("oldest_days") is not None:
        facts.append({"title": "Oldest outstanding", "value": f"{stats['oldest_days']} days"})

    pct = f"{stats['responded'] * 100 // stats['notified']}%" if stats["notified"] else "n/a"
    body += [
        {"type": "TextBlock", "text": "📊 Firewall Rule Recertification — Status",
         "weight": "Bolder", "size": "Medium", "color": "Accent"},
        {"type": "TextBlock", "text": f"Response rate: **{pct}**", "wrap": True},
        {"type": "FactSet", "facts": facts, "spacing": "Medium"},
    ]

    if stats.get("holdouts"):
        lines = "\n\n".join(
            f"• **{c}** — {n} rules" + (f", {d} days" if d is not None else "")
            for c, n, d in stats["holdouts"]
        )
        body += [
            {"type": "TextBlock", "text": "**Largest outstanding owners:**",
             "weight": "Bolder", "spacing": "Medium"},
            {"type": "TextBlock", "text": lines, "wrap": True, "spacing": "Small"},
        ]

    body.append({"type": "TextBlock",
                 "text": f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')} by the NPS Automation team.",
                 "wrap": True, "spacing": "Medium", "isSubtle": True})
    return _card(body)


# ── Teams sending ──────────────────────────────────────────────────────────────

def teams_destination(test_mode=False, per_user=False):
    """Resolve which webhook a card should go to, and a label for logging."""
    if test_mode and TEAMS_WEBHOOK_TEST:
        return TEAMS_WEBHOOK_TEST, "test webhook"
    if per_user and TEAMS_USER_WEBHOOK:
        return TEAMS_USER_WEBHOOK, "per-user flow"
    return TEAMS_WEBHOOK, "channel webhook"


def post_teams(payload, webhook, label, description, dry_run=True, recipient=None):
    """POST one card. recipient is passed through for per-user Power Automate flows."""
    if dry_run:
        log.info(f"[DRY RUN] Would post Teams card ({label}): {description}")
        return True

    if not webhook or webhook.startswith("<"):
        log.warning(f"Teams webhook not configured ({label}), skipping: {description}")
        return False

    if recipient:
        payload = dict(payload, recipient=recipient)

    try:
        resp = requests.post(webhook, headers={"Content-Type": "application/json"},
                             data=json.dumps(payload), timeout=10)
        if resp.status_code in (200, 202):
            log.info(f"Teams card sent ({label}): {description}")
            return True
        log.error(f"Teams post failed ({label}) for {description}: {resp.status_code} {resp.text}")
        return False
    except Exception as e:
        log.error(f"Teams error ({label}) for {description}: {e}")
        return False


def send_teams(owner, rules, dry_run=True, test_mode=False,
               reminder=False, days_outstanding=None):
    """Send one owner-directed card, per-user if a flow is configured."""
    webhook, label = teams_destination(test_mode=test_mode, per_user=True)
    if test_mode and not TEAMS_WEBHOOK_TEST and not dry_run:
        log.warning("TEAMS_WEBHOOK_TEST is empty — test card is going to the real destination")
    if label == "channel webhook" and not test_mode:
        log.warning(f"No per-user flow configured — {owner['corpid']}'s card is going to the channel")

    payload = build_owner_card(owner, rules, test_mode=test_mode,
                               reminder=reminder, days_outstanding=days_outstanding)
    return post_teams(payload, webhook, label,
                      f"{owner['corpid']} | {len(rules)} rules"
                      + (" | reminder" if reminder else ""),
                      dry_run=dry_run, recipient=owner["email"])


def send_teams_channel(payload, description, dry_run=True, test_mode=False):
    """Send one channel-directed card (announce / digest)."""
    webhook, label = teams_destination(test_mode=test_mode, per_user=False)
    return post_teams(payload, webhook, label, description, dry_run=dry_run)


# ── Logging ────────────────────────────────────────────────────────────────────

def write_log(log_file, rows):
    Path(LOGS_DIR).mkdir(exist_ok=True)
    log_file = Path(LOGS_DIR) / log_file
    with open(log_file, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "timestamp", "email", "name", "corpid", "role",
            "rule_count", "attachment_sent", "dry_run", "test_mode",
            "email_sent", "teams_sent", "device_list"
        ])
        writer.writeheader()
        writer.writerows(rows)
    log.info(f"Notification log written to: {log_file}")


def ensure_workspace(log_rows):
    """After a genuine send, create the folders and files used to track replies."""
    if DRY_RUN or TEST_MODE:
        return
    if not any(r["email_sent"] for r in log_rows):
        return

    if not Path(RESPONSES_DIR).exists():
        Path(RESPONSES_DIR).mkdir(exist_ok=True)
        log.info(f"Created {RESPONSES_DIR}/ — save replied-to spreadsheets here")

    Path(LOGS_DIR, "discarded").mkdir(parents=True, exist_ok=True)

    responded = Path(RESPONDED_FILE)
    if not responded.exists():
        responded.write_text("corpid\n", encoding="utf-8")
        log.info(f"Created {RESPONDED_FILE} — add one CorpID per line as replies arrive")


# ── Main ───────────────────────────────────────────────────────────────────────

def notify_run(owner_map, reminder=False, responded=None, history=None):
    """Contact owners: email + optional companion/reminder Teams card."""
    responded = responded or set()
    history   = history or {}
    log_rows  = []

    if TEST_MODE and TEST_LIMIT > 0 and len(owner_map) > TEST_LIMIT:
        log.warning(f"TEST MODE — capping at {TEST_LIMIT} of {len(owner_map)} owner(s). "
                    f"Raise TEST_LIMIT, or set it to 0, to send them all.")
        owner_map = dict(list(owner_map.items())[:TEST_LIMIT])

    for email, data in owner_map.items():
        owner, owner_rules = data["owner"], data["rules"]
        days = days_since(history.get(owner["corpid"].upper()))
        to   = SENDER_EMAIL if TEST_MODE else email
        email_ok = teams_ok = False

        if NOTIFY_EMAIL:
            email_ok = send_email(to, owner, owner_rules, dry_run=DRY_RUN,
                                  reminder=reminder, days_outstanding=days)
        if NOTIFY_TEAMS:
            teams_ok = send_teams(owner, owner_rules, dry_run=DRY_RUN, test_mode=TEST_MODE,
                                  reminder=reminder, days_outstanding=days)
            if not DRY_RUN:
                time.sleep(TEAMS_DELAY_SECONDS)

        devices = list({r["device_name"] for r in owner_rules if r["device_name"]})
        log_rows.append({
            "timestamp":       datetime.now(timezone.utc).isoformat(),
            "email":           email,
            "name":            owner["name"],
            "corpid":          owner["corpid"],
            "role":            owner["role"],
            "rule_count":      len(owner_rules),
            "attachment_sent": True,
            "dry_run":         DRY_RUN,
            "test_mode":       TEST_MODE,
            "email_sent":      email_ok,
            "teams_sent":      teams_ok,
            "device_list":     "; ".join(devices[:5]) + ("..." if len(devices) > 5 else ""),
        })

    return log_rows


def main():
    mode = RUN_MODE.strip().lower()
    if mode not in ("notify", "remind", "announce", "digest"):
        log.error(f'RUN_MODE "{RUN_MODE}" not recognised. '
                  f'Use one of: notify, remind, announce, digest')
        return

    log.info(f"RUN_MODE: {mode}")
    if DRY_RUN:
        log.info("=" * 60)
        log.info("DRY RUN MODE — no notifications will actually be sent")
        log.info("Set DRY_RUN = False in CONFIG to send for real")
        log.info("=" * 60)
    if TEST_MODE and not DRY_RUN:
        log.info(f"TEST MODE — email goes to {SENDER_EMAIL}, Teams cards are labelled as tests")

    rules        = load_rules(EXCEL_PATH)
    owner_map    = build_owner_rule_map(rules)
    total_owners = len(owner_map)

    owner_map = apply_filters(owner_map)
    log.info(f"Unique owners before filters: {total_owners}")
    log.info(f"Unique owners after filters:  {len(owner_map)}")
    log_active_filters()

    if not owner_map:
        log.warning("No owners matched the filters — nothing to do. Check your filter lists.")
        return

    responded = load_responded() if mode in ("remind", "digest") else set()
    history   = load_notification_history() if mode in ("remind", "digest") else {}
    log_rows  = []

    # ── Channel-only modes: no owner is contacted
    if mode == "announce":
        payload = build_announcement_card(len(owner_map), sum(len(d["rules"]) for d in owner_map.values()),
                                          test_mode=TEST_MODE)
        send_teams_channel(payload, f"announcement | {len(owner_map)} owners",
                           dry_run=DRY_RUN, test_mode=TEST_MODE)

    elif mode == "digest":
        stats   = build_stats(owner_map, responded, history)
        payload = build_digest_card(stats, test_mode=TEST_MODE)
        send_teams_channel(payload, f"digest | {stats['outstanding']} outstanding",
                           dry_run=DRY_RUN, test_mode=TEST_MODE)
        log.info(f"Responded: {stats['responded']} | Outstanding: {stats['outstanding']} "
                 f"| Rules outstanding: {stats['rules_outstanding']}")

    # ── Owner-directed modes
    elif mode == "remind":
        before = len(owner_map)
        owner_map = {e: d for e, d in owner_map.items()
                     if d["owner"]["corpid"].upper() not in responded}
        log.info(f"Skipped {before - len(owner_map)} owner(s) already recorded as responded")

        if REMIND_AFTER_DAYS > 0 and history:
            aged = {}
            for e, d in owner_map.items():
                days = days_since(history.get(d["owner"]["corpid"].upper()))
                if days is not None and days >= REMIND_AFTER_DAYS:
                    aged[e] = d
            log.info(f"Skipped {len(owner_map) - len(aged)} owner(s) notified less than "
                     f"{REMIND_AFTER_DAYS} days ago or with no recorded notification")
            owner_map = aged

        if not owner_map:
            log.warning("Nobody is due a reminder right now.")
            return
        log.info(f"Reminding {len(owner_map)} owner(s)")
        log_rows = notify_run(owner_map, reminder=True, responded=responded, history=history)

    else:  # notify
        log_rows = notify_run(owner_map)

    if log_rows:
        write_log(LOG_FILE, log_rows)
        ensure_workspace(log_rows)
        if not DRY_RUN and not TEST_MODE:
            log.info(f"To discard this run, delete {LOGS_DIR}/{LOG_FILE} or move it into "
                     f"{LOGS_DIR}/discarded — it will stop counting as notified")

    print("\n" + "=" * 60)
    print("  RUN SUMMARY")
    print(f"  Run mode:                 {mode}")
    print(f"  Total rules loaded:       {len(rules)}")
    print(f"  Owners found:             {total_owners}")
    print(f"  Owners in scope:          {len(owner_map)}")
    if log_rows:
        print(f"  Owners contacted:         {len(log_rows)}")
        print(f"  Rules in notifications:   {sum(r['rule_count'] for r in log_rows)}")
    if mode in ("remind", "digest"):
        print(f"  Recorded as responded:    {len(responded)}")
    print(f"  NOTIFY_ONLY:              {NOTIFY_ONLY or 'all owners'}")
    print(f"  EXCLUDE_OWNERS:           {EXCLUDE_OWNERS or 'none'}")
    print(f"  NOTIFY_DEVICES:           {NOTIFY_DEVICES or 'all devices'}")
    print(f"  Email notifications:      {'enabled' if NOTIFY_EMAIL else 'disabled'}")
    print(f"  Teams notifications:      {'enabled' if NOTIFY_TEAMS else 'disabled'}")
    print(f"  Test mode:                {TEST_MODE}")
    print(f"  Dry run:                  {DRY_RUN}")
    if log_rows:
        print(f"  Log file:                 {LOGS_DIR}/{LOG_FILE}")
    print("=" * 60)


if __name__ == "__main__":
    main()
